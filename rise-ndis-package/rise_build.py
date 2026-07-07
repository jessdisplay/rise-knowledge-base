#!/usr/bin/env python3
"""Rise document-suite builder: generates rise-document-register.xlsx and rise-nodes-and-edges.json
from one dataset. Verification labels: V=verified vs primary/regulator source Jul 2026;
C=training knowledge corroborated by secondary sources; T=training knowledge, confirm against
current compilation; D=draft instrument, confirm final wording."""
import json, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TODAY = "2026-07-05"

# ---------------- LEGISLATION NODES ----------------
# key, title, kind, register_id/notes, verification, graph role
LEG = [
 ("ACT","National Disability Insurance Scheme Act 2013 (Cth)","Act","Principal Act; amended by Integrity and Safeguarding Act 2026","V","Root authority"),
 ("IS26","National Disability Insurance Scheme Amendment (Integrity and Safeguarding) Act 2026","Amending Act","Passed 1 Apr 2026; Royal Assent 8 Apr 2026. Sch 1 (Commission powers) and Sch 3 (whistleblower) commenced 9 Apr 2026; Sch 2 (NDIA operational incl. electronic claims) 6 May 2026. Provider effects: serious-contravention civil penalties; offence to provide registration-required supports while unregistered; expanded banning orders incl. auditors and advisors; anti-promotion orders; stronger whistleblower protections","V","Amends ACT"),
 ("PRPS","NDIS (Provider Registration and Practice Standards) Rules 2018","Rules (delegated legislation)","FRL series F2018L00631. Practice Standards housed in Schedules 1-8 (Core = Sch 1; HIDPA = Sch 2; M2 = Sch 3; M2A = Sch 4; M3 = Sch 5; M4 = Sch 6; M5 SDA = Sch 7; Verification 'Module 6' = Sch 8). Body: s 7 registration classes, s 13B mid-term audit, s 20 applicable standards, s 24 QI Guidelines link. Latest published compilation C04 = F2021C01137 (15 Nov 2021) as at 5 Jul 2026 - the 2026 SIL amendments are made but not yet consolidated at this URL","V","Practice Standards PUBLISHED_IN here; MADE_UNDER ACT (registration power incl. s 73E(1))"),
 ("PRPS26","NDIS (Provider Registration and Practice Standards) Amendment (Mandatory Registration and Other Matters) Rules 2026","Amending Rules","Mandatory SIL & platform provider registration + SIL Practice Standards, from 1 Jul 2026","V","Amends PRPS"),
 ("QI","NDIS (Quality Indicators for NDIS Practice Standards) Guidelines 2018","Guidelines","F2018N00041; made under s 181D(2) of the Act; registered 18 May 2018, commenced 1 Jul 2018. Renamed from (Quality Indicators) Guidelines 2018. Current compilation C03 = F2026C00528, in force 1 Jul 2026 - includes the Module 5A (SIL) indicators at ss 72B-72E","V","One section of indicators per standard; auditors assess against these"),
 ("AQA","NDIS (Approved Quality Auditors Scheme) Guidelines 2018","Guidelines (notifiable instrument)","F2018N00114. Governs certification and verification audits; audit reports due to Commission 14 days (verification) / 28 days (certification, mid-term) after completion. New AQA approvals suspended pending 2025-26 reforms; see also AQAR25","V","Governs the Audit node lifecycle"),
 ("AQAR25","NDIS (Approved Quality Auditors) Rules 2025","Rules","F2025L01383, listed on the Commission legislation page (May 2026); relationship to the 2018 AQA Scheme Guidelines to be confirmed","V","Audit scheme (newer instrument)"),
 ("QI26","NDIS (Quality Indicators for NDIS Practice Standards) Amendment (Supported Independent Living) Guidelines 2026","Amending Guidelines","SIL quality indicators; consolidated into the principal Guidelines at compilation C03 (1 Jul 2026)","V","Amends QI"),
 ("IMRI","NDIS (Incident Management and Reportable Incidents) Rules 2018","Rules","F2018L00633; incident systems + reportable incident notification","V","MADE_UNDER ACT"),
 ("CMR","NDIS (Complaints Management and Resolution) Rules 2018","Rules","F2018L00634; complaints systems + records","V","MADE_UNDER ACT"),
 ("RPBS","NDIS (Restrictive Practices and Behaviour Support) Rules 2018","Rules","Series F2018L00632 (compilations incl. F2020C01087); regulated restrictive practices, BSPs, monthly RP reporting","V","MADE_UNDER ACT"),
 ("WSR","NDIS (Practice Standards - Worker Screening) Rules 2018","Rules","F2018L00887. Worker screening obligations for risk-assessed roles","V","MADE_UNDER ACT"),
 ("COC","NDIS (Code of Conduct) Rules 2018","Rules","F2018L00629. NDIS Code of Conduct for providers and workers","V","MADE_UNDER ACT"),
 ("SDAC","NDIS (Specialist Disability Accommodation Conditions) Rule 2018","Rule","Series F2018L00627 (compilations incl. F2020C00549); SDA conditions","V","MADE_UNDER ACT"),
 ("PD","NDIS (Provider Definition) Rule 2018","Rule","Series F2018L00628 (compilations incl. F2021C00694)","V","MADE_UNDER ACT"),
 ("NCE","NDIS (Registered NDIS Provider Notice of Changes and Events) Guidelines 2019","Guidelines","Listed on the Commission legislation page (May 2026); FRL id TBC. Read with PRPS rr 13-13A; PRPS26 shortens notification timeframes and strengthens change-of-ownership notice (per Commission summary of the amendments)","V","Guides r 13/13A notifications"),
 ("BSPA","NDIS (NDIS Behaviour Support Practitioner Application) Guidelines 2020","Guidelines","Listed on the Commission legislation page (May 2026); FRL id TBC. Practitioner suitability process for behaviour support","V","Applies to M2 specialist providers"),
 ("STWS","State/Territory NDIS worker screening laws (jurisdictional)","Act (adjacent)","NDIS Worker Screening Check is issued under state/territory schemes - identify applicable jurisdiction(s)","C","Adjacent law referenced by documents"),
 ("DDA","Disability Discrimination Act 1992 (Cth)","Act (adjacent)","Anti-discrimination duties relevant to rights standards","C","Adjacent law referenced by documents"),
 ("PRIV","Privacy Act 1988 (Cth) incl. Australian Privacy Principles & Notifiable Data Breaches scheme","Act (adjacent)","Applies to handling of personal information","C","Adjacent law referenced by documents"),
 ("WHSL","Work Health and Safety Act (jurisdictional)","Act (adjacent)","State/Territory WHS law; identify applicable jurisdiction","C","Adjacent law referenced by documents"),
]

# ---------------- FRAMEWORK: MODULES + STANDARDS ----------------
MODULES = [
 ("CORE","Core Module","All registered providers of higher-risk supports (certification pathway)","V"),
 ("VER","Module 6: Verification (Sch 8)","Providers of lower-risk supports (verification pathway); legislative label is Module 6","V"),
 ("M1","Module 1: High Intensity Daily Personal Activities (Sch 2)","Providers delivering HIDPA supports; read with HIDPA skills descriptors","V"),
 ("M2","Module 2: Specialist Behaviour Support (Sch 3)","Specialist behaviour support providers (developing BSPs)","V"),
 ("M2A","Module 2A: Implementing Behaviour Support Plans (Sch 4)","Providers implementing BSPs incl. regulated restrictive practices","V"),
 ("M3","Module 3: Early Childhood Supports","Providers of early childhood supports","V"),
 ("M4","Module 4: Specialised Support Coordination","Providers of specialised support coordination","V"),
 ("M5","Module 5: Specialist Disability Accommodation","SDA providers","V"),
 ("SIL","Module 5A: Assistance with Supported Independent Living (SIL)","SIL providers; commenced 1 Jul 2026 with mandatory registration; official module name per QI Guidelines C03, Part 8A","V"),
]
# std_id, module, division, name, verification, citation (PRPS Rules, per compilation F2020C00051;
# 2021/2026 insertions marked TBC pending check of the latest compilation)
STD = [
 # Core Division 1
 ("CORE-1.1","CORE","Sch 1 Pt 2 - Rights of participants & responsibilities of providers","Person-centred supports","V"),
 ("CORE-1.2","CORE","Sch 1 Pt 2 - Rights of participants & responsibilities of providers","Individual values and beliefs","V"),
 ("CORE-1.3","CORE","Sch 1 Pt 2 - Rights of participants & responsibilities of providers","Privacy and dignity","V"),
 ("CORE-1.4","CORE","Sch 1 Pt 2 - Rights of participants & responsibilities of providers","Independence and informed choice","V"),
 ("CORE-1.5","CORE","Sch 1 Pt 2 - Rights of participants & responsibilities of providers","Freedom from violence, abuse, neglect, exploitation or discrimination","V"),
 # Core Division 2
 ("CORE-2.1","CORE","Sch 1 Pt 3 - Provider governance & operational management","Governance and operational management","V"),
 ("CORE-2.2","CORE","Sch 1 Pt 3 - Provider governance & operational management","Risk management","V"),
 ("CORE-2.3","CORE","Sch 1 Pt 3 - Provider governance & operational management","Quality management","V"),
 ("CORE-2.4","CORE","Sch 1 Pt 3 - Provider governance & operational management","Information management","V"),
 ("CORE-2.5","CORE","Sch 1 Pt 3 - Provider governance & operational management","Complaints management and resolution","V"),
 ("CORE-2.6","CORE","Sch 1 Pt 3 - Provider governance & operational management","Incident management","V"),
 ("CORE-2.7","CORE","Sch 1 Pt 3 - Provider governance & operational management","Human resource management","V"),
 ("CORE-2.8","CORE","Sch 1 Pt 3 - Provider governance & operational management","Continuity of supports","V"),
 ("CORE-2.9","CORE","Sch 1 Pt 3 - Provider governance & operational management","Emergency and disaster management (added 2021)","V"),
 # Core Division 3
 ("CORE-3.1","CORE","Sch 1 Pt 4 - Provision of supports","Access to supports","V"),
 ("CORE-3.2","CORE","Sch 1 Pt 4 - Provision of supports","Support planning","V"),
 ("CORE-3.3","CORE","Sch 1 Pt 4 - Provision of supports","Service agreements","V"),
 ("CORE-3.4","CORE","Sch 1 Pt 4 - Provision of supports","Responsive support provision","V"),
 ("CORE-3.5","CORE","Sch 1 Pt 4 - Provision of supports","Transitions to and from a provider","V"),
 # Core Division 4
 ("CORE-4.1","CORE","Sch 1 Pt 5 - Support provision environment","Safe environment","V"),
 ("CORE-4.2","CORE","Sch 1 Pt 5 - Support provision environment","Participant money and property","V"),
 ("CORE-4.3","CORE","Sch 1 Pt 5 - Support provision environment","Management of medication","V"),
 ("CORE-4.4","CORE","Sch 1 Pt 5 - Support provision environment","Mealtime management (added 2021)","V"),
 ("CORE-4.5","CORE","Sch 1 Pt 5 - Support provision environment","Management of waste","V"),
 # Verification module (verified against Commission website 5 Jul 2026)
 ("VER-1","VER","Verification","Human resource management","V"),
 ("VER-2","VER","Verification","Incident management","V"),
 ("VER-3","VER","Verification","Complaints management and resolution","V"),
 ("VER-4","VER","Verification","Risk management","V"),
 # Module 1 HIDPA skills areas
 ("M1-1","M1","HIDPA","Complex bowel care","V"),
 ("M1-2","M1","HIDPA","Enteral (naso-gastric tube-jejunum or duodenum) feeding and management","V"),
 ("M1-3","M1","HIDPA","Severe dysphagia management (added 2021)","V"),
 ("M1-4","M1","HIDPA","Tracheostomy management","V"),
 ("M1-5","M1","HIDPA","Urinary catheter management","V"),
 ("M1-6","M1","HIDPA","Ventilator management","V"),
 ("M1-7","M1","HIDPA","Subcutaneous injections","V"),
 ("M1-8","M1","HIDPA","Complex wound management","V"),
 # Module 2 (Sch 3) - specialist behaviour support providers
 ("M2-1","M2","Specialist behaviour support","Behaviour support in the NDIS","V"),
 ("M2-2","M2","Specialist behaviour support","Regulated restrictive practices","V"),
 ("M2-3","M2","Specialist behaviour support","Behaviour support plans","V"),
 ("M2-4","M2","Specialist behaviour support","Supporting the implementation of a behaviour support plan","V"),
 ("M2-5","M2","Specialist behaviour support","Behaviour support plan monitoring and review","V"),
 ("M2-6","M2","Specialist behaviour support","Reportable incidents involving the use of a regulated restrictive practice","V"),
 ("M2-7","M2","Specialist behaviour support","Interim behaviour support plans","V"),
 # Module 2A (Sch 4) - providers implementing BSPs / using regulated restrictive practices
 ("M2A-1","M2A","Implementing behaviour support plans","Behaviour support in the NDIS","V"),
 ("M2A-2","M2A","Implementing behaviour support plans","Regulated restrictive practices","V"),
 ("M2A-3","M2A","Implementing behaviour support plans","Supporting the assessment and development of behaviour support plans","V"),
 ("M2A-4","M2A","Implementing behaviour support plans","Supporting the implementation of a behaviour support plan","V"),
 ("M2A-5","M2A","Implementing behaviour support plans","Monitoring and reporting the use of regulated restrictive practices","V"),
 ("M2A-6","M2A","Implementing behaviour support plans","Behaviour support plan monitoring and review","V"),
 ("M2A-7","M2A","Implementing behaviour support plans","Reportable incidents involving the use of a regulated restrictive practice","V"),
 ("M2A-8","M2A","Implementing behaviour support plans","Interim behaviour support plans","V"),
 # Module 3 (training knowledge - confirm)
 ("M3-1","M3","Early childhood","The child","V"),
 ("M3-2","M3","Early childhood","The family","V"),
 ("M3-3","M3","Early childhood","Inclusion","V"),
 ("M3-4","M3","Early childhood","Collaboration","V"),
 ("M3-5","M3","Early childhood","Capacity building","V"),
 ("M3-6","M3","Early childhood","Evidence-informed supports","V"),
 ("M3-7","M3","Early childhood","Outcome based approach","V"),
 # Module 4
 ("M4-1","M4","Support coordination","Specialised support coordination","V"),
 ("M4-2","M4","Support coordination","Management of supports","V"),
 ("M4-3","M4","Support coordination","Conflict of interest","V"),
 # Module 5 SDA
 ("M5-1","M5","SDA","Rights and responsibilities","V"),
 ("M5-2","M5","SDA","Conflict of interest","V"),
 ("M5-3","M5","SDA","Service agreements with participants","V"),
 ("M5-4","M5","SDA","Enrolment of SDA dwellings","V"),
 ("M5-5","M5","SDA","Tenancy management","V"),
 # SIL module (draft published by Commission; commences 1 Jul 2026)
 ("SIL-1","SIL","Module 5A (SIL)","Supported decision-making","V"),
 ("SIL-2","SIL","Module 5A (SIL)","Safeguarding","V"),
 ("SIL-3","SIL","Module 5A (SIL)","Practice governance","V"),
 ("SIL-4","SIL","Module 5A (SIL)","Agreements about tenancy, housing and support arrangements","V"),
]

# Clause citations, NDIS (Provider Registration and Practice Standards) Rules 2018 (F2018L00631),
# read from compilation F2020C00051 ToC. Amending-instrument insertions marked TBC.
TBC21 = "inserted by 2021 amendments - clause no. TBC vs latest compilation"
CIT = {}
for n,i in enumerate(["1.1","1.2","1.3","1.4","1.5"]): CIT["CORE-"+i] = f"Sch 1, Pt 2, cl {n+3}"
for n,i in enumerate(["2.1","2.2","2.3","2.4","2.5","2.6","2.7","2.8"]): CIT["CORE-"+i] = f"Sch 1, Pt 3, cl {n+9}"
CIT["CORE-2.9"] = "Sch 1, Pt 3, cl 16A"
for n,i in enumerate(["3.1","3.2","3.3","3.4","3.5"]): CIT["CORE-"+i] = f"Sch 1, Pt 4, cl {n+18}"
for n,i in enumerate(["4.1","4.2","4.3"]): CIT["CORE-"+i] = f"Sch 1, Pt 5, cl {n+24}"
CIT["CORE-4.4"] = "Sch 1, Pt 5, cl 26A"
CIT["CORE-4.5"] = "Sch 1, Pt 5, cl 27"
for n,i in enumerate([1,2]): CIT[f"M1-{i}"] = f"Sch 2, cl {n+3}"
CIT["M1-3"] = "Sch 2, cl 4A"
for n,i in enumerate([4,5,6,7,8]): CIT[f"M1-{i}"] = f"Sch 2, cl {n+5}"
for i in range(1,8): CIT[f"M2-{i}"] = f"Sch 3, cl {i+2}"
for i in range(1,9): CIT[f"M2A-{i}"] = f"Sch 4, cl {i+2}"
for i in range(1,8): CIT[f"M3-{i}"] = f"Sch 5, cl {i+2}"
for i in range(1,4): CIT[f"M4-{i}"] = f"Sch 6, cl {i+2}"
for i in range(1,6): CIT[f"M5-{i}"] = f"Sch 7, cl {i+2}"
CIT.update({"VER-1":"Sch 8, cl 6","VER-2":"Sch 8, cl 5","VER-3":"Sch 8, cl 4","VER-4":"Sch 8, cl 3"})
for i in range(1,5): CIT[f"SIL-{i}"] = "Inserted by 2026 Amendment Rules - Sch/cl TBC (check latest compilation)"

# ---- Quality indicator section per standard (QI Guidelines 2018, compilation C03, 1 Jul 2026; ToC verified 6 Jul 2026) ----
QIREF = {}
for i, n in enumerate(["6","7","8","9","10"]): QIREF[f"CORE-1.{i+1}"] = "s " + n
for i, n in enumerate(["11","12","13","14","15","16","17","18","18A"]): QIREF[f"CORE-2.{i+1}"] = "s " + n
for i, n in enumerate(["19","20","21","22","23"]): QIREF[f"CORE-3.{i+1}"] = "s " + n
for i, n in enumerate(["24","25","26","26A","27"]): QIREF[f"CORE-4.{i+1}"] = "s " + n
for i, n in enumerate(["29","30","30A","31","32","33","34","35"]): QIREF[f"M1-{i+1}"] = "s " + n
for i in range(7): QIREF[f"M2-{i+1}"] = "s " + str(38+i)
for i in range(8): QIREF[f"M2A-{i+1}"] = "s " + str(47+i)
for i in range(7): QIREF[f"M3-{i+1}"] = "s " + str(56+i)
for i in range(3): QIREF[f"M4-{i+1}"] = "s " + str(64+i)
for i in range(5): QIREF[f"M5-{i+1}"] = "s " + str(68+i)
for i, n in enumerate(["72B","72C","72D","72E"]): QIREF[f"SIL-{i+1}"] = "s " + n
for i, n in enumerate(["74","75","76","77"]): QIREF[f"VER-{i+1}"] = "s " + n
QIMOD = {"CORE":"ss 5, 11A, 19A, 24A","M1":"s 28","M2":"ss 36-37","M2A":"ss 45-46","M3":"s 55","M4":"s 63","M5":"s 67","SIL":"s 72A","VER":"s 73"}

# ---------------- DOCUMENT SUITE ----------------
# (id, type, title, cluster, applies, priority, implements[], legislation keys[], easy_read, owner, review_yrs)
def d(i,t,ti,cl,ap,pr,imp,leg,ez=0,own="Quality Manager",rv=2):
    return dict(id=i,type=t,title=ti,cluster=cl,applies=ap,priority=pr,imp=imp,leg=leg,ez=ez,own=own,rv=rv)
DOCS = [
 # Governance
 d("POL-GOV-01","Policy","Governance and Operational Management Policy","Governance","Core","P1",["CORE-2.1"],["PRPS","ACT"],0,"Board / CEO"),
 d("PRO-GOV-01","Procedure","Delegations and Decision-Making Procedure","Governance","Core","P2",["CORE-2.1"],["PRPS"],0,"CEO"),
 d("REG-GOV-01","Register","Delegations Register","Governance","Core","P2",["CORE-2.1"],["PRPS"],0,"CEO",1),
 d("POL-GOV-02","Policy","Conflict of Interest Policy","Governance","Core; M4; M5","P1",["CORE-2.1","M4-3","M5-2"],["PRPS"],0,"Board / CEO"),
 d("FRM-GOV-01","Form","Conflict of Interest Declaration Form","Governance","Core","P2",["CORE-2.1"],["PRPS"],0,"CEO",1),
 d("REG-GOV-02","Register","Conflict of Interest Register","Governance","Core","P2",["CORE-2.1"],["PRPS"],0,"CEO",1),
 d("REG-GOV-03","Register","Master Document Register (this workbook)","Governance","All","P1",["CORE-2.3"],["PRPS"],0,"Quality Manager",1),
 d("POL-GOV-03","Policy","Whistleblower and Disclosure Protection Policy","Governance","All","P1",["CORE-2.1","CORE-2.5"],["IS26","ACT","COC"],0,"Board / CEO"),
 d("PRO-GOV-02","Procedure","Notification of Changes and Events Procedure (Commission notice under rr 13-13A)","Governance","All","P1",["CORE-2.1"],["PRPS","PRPS26","NCE"],0,"CEO"),
 # Risk
 d("POL-RSK-01","Policy","Risk Management Policy","Risk","Core; Verification","P1",["CORE-2.2","VER-4"],["PRPS"]),
 d("PRO-RSK-01","Procedure","Risk Assessment and Treatment Procedure","Risk","Core; Verification","P1",["CORE-2.2","VER-4"],["PRPS"]),
 d("FRM-RSK-01","Form","Risk Assessment Form","Risk","Core; Verification","P1",["CORE-2.2"],["PRPS"],0,"Quality Manager",1),
 d("REG-RSK-01","Register","Risk Register","Risk","Core; Verification","P1",["CORE-2.2","VER-4"],["PRPS"],0,"Quality Manager",1),
 # Quality
 d("POL-QMS-01","Policy","Quality Management and Continuous Improvement Policy","Quality","Core","P1",["CORE-2.3"],["PRPS"]),
 d("PRO-QMS-01","Procedure","Continuous Improvement Procedure","Quality","Core","P2",["CORE-2.3"],["PRPS"]),
 d("PRO-QMS-02","Procedure","Document and Records Control Procedure","Quality","Core","P1",["CORE-2.3","CORE-2.4"],["PRPS"]),
 d("PRO-QMS-03","Procedure","Internal Audit and Self-Assessment Procedure","Quality","Core","P2",["CORE-2.3"],["PRPS","QI"]),
 d("FRM-QMS-01","Form","Improvement / Corrective Action Request Form","Quality","Core","P2",["CORE-2.3"],["PRPS"],0,"Quality Manager",1),
 d("REG-QMS-01","Register","Continuous Improvement Register","Quality","Core","P1",["CORE-2.3"],["PRPS"],0,"Quality Manager",1),
 d("REG-QMS-02","Register","Internal Audit Schedule and Register","Quality","Core","P2",["CORE-2.3"],["PRPS","QI"],0,"Quality Manager",1),
 # Information & privacy
 d("POL-INF-01","Policy","Information Management and Privacy Policy","Information","Core","P1",["CORE-2.4","CORE-1.3"],["PRPS","PRIV"]),
 d("PRO-INF-01","Procedure","Records Management Procedure","Information","Core","P2",["CORE-2.4"],["PRPS","PRIV"]),
 d("PRO-INF-02","Procedure","Privacy and Data Breach Response Procedure","Information","Core","P2",["CORE-2.4"],["PRIV"]),
 d("FRM-INF-01","Form","Consent to Collect, Use and Share Information Form","Information","Core","P1",["CORE-2.4","CORE-1.4"],["PRPS","PRIV"],1,"Quality Manager",1),
 d("REG-INF-01","Register","Data Breach Register","Information","Core","P3",["CORE-2.4"],["PRIV"],0,"Quality Manager",1),
 # Feedback & complaints
 d("POL-FBK-01","Policy","Feedback and Complaints Policy","Complaints","Core; Verification","P1",["CORE-2.5","VER-3"],["PRPS","CMR"],1),
 d("PRO-FBK-01","Procedure","Complaints Handling and Resolution Procedure","Complaints","Core; Verification","P1",["CORE-2.5","VER-3"],["CMR"]),
 d("WIN-FBK-01","Work Instruction","Receiving and Recording a Complaint (frontline)","Complaints","Core; Verification","P2",["CORE-2.5"],["CMR"],0,"Service Manager",1),
 d("FRM-FBK-01","Form","Feedback and Complaints Form","Complaints","Core; Verification","P1",["CORE-2.5","VER-3"],["CMR"],1,"Quality Manager",1),
 d("REG-FBK-01","Register","Complaints Register","Complaints","Core; Verification","P1",["CORE-2.5","VER-3"],["CMR"],0,"Quality Manager",1),
 # Incidents
 d("POL-INC-01","Policy","Incident Management Policy","Incidents","Core; Verification","P1",["CORE-2.6","VER-2"],["PRPS","IMRI"]),
 d("PRO-INC-01","Procedure","Incident Management Procedure","Incidents","Core; Verification","P1",["CORE-2.6","VER-2"],["IMRI"]),
 d("PRO-INC-02","Procedure","Reportable Incidents Notification Procedure","Incidents","Core; Verification","P1",["CORE-2.6","VER-2"],["IMRI"]),
 d("WIN-INC-01","Work Instruction","Responding to an Incident (frontline)","Incidents","Core; Verification","P1",["CORE-2.6"],["IMRI"],0,"Service Manager",1),
 d("FRM-INC-01","Form","Incident Report Form","Incidents","Core; Verification","P1",["CORE-2.6","VER-2"],["IMRI"],0,"Quality Manager",1),
 d("REG-INC-01","Register","Incident Register","Incidents","Core; Verification","P1",["CORE-2.6","VER-2"],["IMRI"],0,"Quality Manager",1),
 # Human resources
 d("POL-HRM-01","Policy","Human Resource Management Policy","People","Core; Verification","P1",["CORE-2.7","VER-1"],["PRPS","WSR"]),
 d("PRO-HRM-01","Procedure","Recruitment, Screening and Induction Procedure","People","Core; Verification","P1",["CORE-2.7","VER-1"],["WSR","COC"],0,"HR Manager"),
 d("PRO-HRM-02","Procedure","Supervision and Performance Procedure","People","Core","P2",["CORE-2.7"],["PRPS"],0,"HR Manager"),
 d("PRO-HRM-03","Procedure","Training and Competency Procedure","People","Core","P1",["CORE-2.7"],["PRPS"],0,"HR Manager"),
 d("FRM-HRM-01","Form","Worker Screening and Credential Checklist","People","Core; Verification","P1",["CORE-2.7","VER-1"],["WSR"],0,"HR Manager",1),
 d("FRM-HRM-02","Form","NDIS Code of Conduct Acknowledgement Form","People","All","P1",["CORE-2.7"],["COC"],0,"HR Manager",1),
 d("REG-HRM-01","Register","Worker Screening Register","People","Core; Verification","P1",["CORE-2.7","VER-1"],["WSR"],0,"HR Manager",1),
 d("REG-HRM-02","Register","Training and Competency Register","People","Core","P1",["CORE-2.7"],["PRPS"],0,"HR Manager",1),
 d("POL-COC-01","Policy","NDIS Code of Conduct Policy","People","All","P1",["CORE-2.7","CORE-1.5"],["COC"],1,"CEO"),
 # Continuity
 d("POL-COS-01","Policy","Continuity of Supports Policy","Continuity","Core","P2",["CORE-2.8"],["PRPS"]),
 d("PLN-COS-01","Plan","Business Continuity Plan","Continuity","Core","P2",["CORE-2.8"],["PRPS"],0,"CEO",1),
 d("PRO-COS-01","Procedure","Worker Absence and Backup Arrangements Procedure","Continuity","Core","P3",["CORE-2.8"],["PRPS"],0,"Service Manager"),
 # Emergency & disaster
 d("POL-EDM-01","Policy","Emergency and Disaster Management Policy","Emergency","Core","P2",["CORE-2.9"],["PRPS"]),
 d("PLN-EDM-01","Plan","Emergency and Disaster Management Plan","Emergency","Core","P2",["CORE-2.9"],["PRPS"],1,"Service Manager",1),
 d("PRO-EDM-01","Procedure","Emergency Response and Evacuation Procedure","Emergency","Core","P2",["CORE-2.9"],["PRPS","WHSL"],0,"Service Manager",1),
 d("REG-EDM-01","Register","Emergency Preparedness and Drills Register","Emergency","Core","P3",["CORE-2.9"],["PRPS"],0,"Service Manager",1),
 # Rights & safeguarding
 d("POL-RGT-01","Policy","Person-Centred Supports Policy","Rights","Core","P1",["CORE-1.1","CORE-1.2"],["PRPS","ACT"],1),
 d("POL-RGT-02","Policy","Privacy and Dignity Policy","Rights","Core","P1",["CORE-1.3"],["PRPS","PRIV"],1),
 d("POL-RGT-03","Policy","Independence, Informed Choice and Supported Decision-Making Policy","Rights","Core; SIL","P1",["CORE-1.4","SIL-1"],["PRPS","PRPS26"],1),
 d("FRM-RGT-01","Form","Communication, Advocacy and Support Preferences Form","Rights","Core","P2",["CORE-1.1","CORE-1.4"],["PRPS"],1,"Service Manager",1),
 d("STA-RGT-01","Handbook","Participant Handbook / Welcome Pack","Rights","Core","P1",["CORE-1.1","CORE-1.4","CORE-3.3"],["PRPS","COC"],1,"Service Manager"),
 d("POL-SGD-01","Policy","Safeguarding Policy (zero tolerance of violence, abuse, neglect, exploitation and discrimination)","Safeguarding","Core; SIL","P1",["CORE-1.5","SIL-2"],["PRPS","COC","IMRI"],1,"CEO"),
 d("PRO-SGD-01","Procedure","Responding to Abuse, Neglect and Exploitation Procedure","Safeguarding","Core; SIL","P1",["CORE-1.5","SIL-2"],["IMRI","COC"]),
 # Provision of supports
 d("POL-SUP-01","Policy","Access to Supports Policy","Supports","Core","P2",["CORE-3.1"],["PRPS"],1),
 d("PRO-SUP-01","Procedure","Intake, Eligibility and Waitlist Procedure","Supports","Core","P2",["CORE-3.1"],["PRPS"],0,"Service Manager"),
 d("FRM-SUP-01","Form","Referral and Intake Form","Supports","Core","P2",["CORE-3.1"],["PRPS"],0,"Service Manager",1),
 d("POL-SUP-02","Policy","Support Planning Policy","Supports","Core","P1",["CORE-3.2"],["PRPS"],1),
 d("PRO-SUP-02","Procedure","Support Planning and Review Procedure","Supports","Core","P1",["CORE-3.2"],["PRPS"],0,"Service Manager"),
 d("FRM-SUP-02","Form","Participant Support Plan Template","Supports","Core","P1",["CORE-3.2"],["PRPS"],1,"Service Manager",1),
 d("POL-SUP-03","Policy","Service Agreements Policy","Supports","Core","P1",["CORE-3.3"],["PRPS"],1),
 d("AGR-SUP-01","Agreement","Service Agreement Template","Supports","Core","P1",["CORE-3.3"],["PRPS"],1,"Service Manager",1),
 d("PRO-SUP-03","Procedure","Responsive Support Provision Procedure","Supports","Core","P2",["CORE-3.4"],["PRPS"],0,"Service Manager"),
 d("POL-SUP-04","Policy","Transitions To and From the Provider Policy","Supports","Core","P2",["CORE-3.5"],["PRPS"]),
 d("PRO-SUP-04","Procedure","Entry, Exit and Transition Procedure","Supports","Core","P2",["CORE-3.5"],["PRPS"],0,"Service Manager"),
 d("FRM-SUP-04","Form","Transition and Exit Checklist","Supports","Core","P3",["CORE-3.5"],["PRPS"],0,"Service Manager",1),
 d("REG-SUP-01","Register","Participant and Service Agreement Register","Supports","Core","P2",["CORE-3.3","CORE-3.1"],["PRPS"],0,"Service Manager",1),
 # Support environment
 d("POL-ENV-01","Policy","Safe Environment Policy","Environment","Core","P1",["CORE-4.1"],["PRPS","WHSL"]),
 d("PRO-ENV-01","Procedure","Home and Workplace Safety Check Procedure","Environment","Core","P2",["CORE-4.1"],["PRPS","WHSL"],0,"Service Manager",1),
 d("FRM-ENV-01","Form","Environmental Safety Checklist","Environment","Core","P2",["CORE-4.1"],["WHSL"],0,"Service Manager",1),
 d("REG-ENV-01","Register","Hazard and Maintenance Register","Environment","Core","P2",["CORE-4.1"],["WHSL"],0,"Service Manager",1),
 d("POL-MNY-01","Policy","Participant Money and Property Policy","Environment","Core","P1",["CORE-4.2"],["PRPS"],1),
 d("PRO-MNY-01","Procedure","Handling Participant Money and Property Procedure","Environment","Core","P1",["CORE-4.2"],["PRPS"],0,"Service Manager"),
 d("FRM-MNY-01","Form","Participant Money Transaction Record","Environment","Core","P1",["CORE-4.2"],["PRPS"],0,"Service Manager",1),
 d("REG-MNY-01","Register","Participant Money and Property Register","Environment","Core","P1",["CORE-4.2"],["PRPS"],0,"Service Manager",1),
 d("POL-MED-01","Policy","Medication Management Policy","Environment","Core","P1",["CORE-4.3"],["PRPS"]),
 d("PRO-MED-01","Procedure","Medication Administration Procedure","Environment","Core","P1",["CORE-4.3"],["PRPS"],0,"Service Manager"),
 d("WIN-MED-01","Work Instruction","Administering PRN Medication (frontline)","Environment","Core","P2",["CORE-4.3"],["PRPS"],0,"Service Manager",1),
 d("FRM-MED-01","Form","Medication Administration Record (chart)","Environment","Core","P1",["CORE-4.3"],["PRPS"],0,"Service Manager",1),
 d("POL-MTM-01","Policy","Mealtime Management Policy","Environment","Core","P2",["CORE-4.4"],["PRPS"]),
 d("PRO-MTM-01","Procedure","Mealtime Management Procedure","Environment","Core","P2",["CORE-4.4","M1-3"],["PRPS"],0,"Service Manager"),
 d("FRM-MTM-01","Form","Mealtime Management Plan","Environment","Core","P2",["CORE-4.4"],["PRPS"],1,"Service Manager",1),
 d("POL-WST-01","Policy","Waste Management and Infection Control Policy","Environment","Core","P3",["CORE-4.5"],["PRPS","WHSL"]),
 d("PRO-WST-01","Procedure","Infection Control and Waste Handling Procedure","Environment","Core","P3",["CORE-4.5"],["WHSL"],0,"Service Manager",1),
 # Module 1 HIDPA
 d("POL-HID-01","Policy","High Intensity Supports Policy","High intensity","M1","P1",["M1-1","M1-2","M1-3","M1-4","M1-5","M1-6","M1-7","M1-8"],["PRPS","QI"]),
 d("PRO-HID-01","Procedure","Complex Bowel Care Procedure","High intensity","M1","P1",["M1-1"],["PRPS"],0,"Clinical Lead",1),
 d("PRO-HID-02","Procedure","Enteral Feeding and Management Procedure","High intensity","M1","P1",["M1-2"],["PRPS"],0,"Clinical Lead",1),
 d("PRO-HID-03","Procedure","Severe Dysphagia Management Procedure","High intensity","M1","P1",["M1-3"],["PRPS"],0,"Clinical Lead",1),
 d("PRO-HID-04","Procedure","Tracheostomy Management Procedure","High intensity","M1","P1",["M1-4"],["PRPS"],0,"Clinical Lead",1),
 d("PRO-HID-05","Procedure","Urinary Catheter Management Procedure","High intensity","M1","P1",["M1-5"],["PRPS"],0,"Clinical Lead",1),
 d("PRO-HID-06","Procedure","Ventilator Management Procedure","High intensity","M1","P1",["M1-6"],["PRPS"],0,"Clinical Lead",1),
 d("PRO-HID-07","Procedure","Subcutaneous Injections Procedure","High intensity","M1","P1",["M1-7"],["PRPS"],0,"Clinical Lead",1),
 d("PRO-HID-08","Procedure","Complex Wound Management Procedure","High intensity","M1","P1",["M1-8"],["PRPS"],0,"Clinical Lead",1),
 d("FRM-HID-01","Form","Participant Health Support Plan","High intensity","M1","P1",["M1-1"],["PRPS"],0,"Clinical Lead",1),
 d("REG-HID-01","Register","High Intensity Training and Competency Register","High intensity","M1","P1",["M1-1"],["PRPS","QI"],0,"Clinical Lead",1),
 # Modules 2 / 2A
 d("POL-BSP-01","Policy","Positive Behaviour Support Policy","Behaviour support","M2; M2A","P1",["M2A-1","M2-1"],["RPBS","PRPS"]),
 d("PRO-BSP-01","Procedure","Behaviour Support Plan Implementation and Monitoring Procedure","Behaviour support","M2A","P1",["M2A-4","M2A-6","M2A-8"],["RPBS"],0,"Service Manager"),
 d("PRO-BSP-02","Procedure","Regulated Restrictive Practices Authorisation, Use and Reporting Procedure","Behaviour support","M2A","P1",["M2A-2","M2A-5","M2A-7"],["RPBS","IMRI"],0,"Service Manager"),
 d("PRO-BSP-03","Procedure","Behaviour Support Plan Development and Review Procedure (specialist providers)","Behaviour support","M2","P2",["M2-3","M2-5","M2-7"],["RPBS"],0,"Clinical Lead"),
 d("FRM-BSP-01","Form","Restrictive Practice Use Record","Behaviour support","M2A","P1",["M2A-5","M2A-7"],["RPBS"],0,"Service Manager",1),
 d("REG-BSP-01","Register","Restrictive Practices Register","Behaviour support","M2A","P1",["M2A-2","M2A-5"],["RPBS"],0,"Service Manager",1),
 # Module 3
 d("POL-ECS-01","Policy","Early Childhood Supports Policy","Early childhood","M3","P2",["M3-1","M3-2","M3-3","M3-4","M3-5","M3-6","M3-7"],["PRPS"]),
 d("PRO-ECS-01","Procedure","Family-Centred Practice and Key Worker Procedure","Early childhood","M3","P2",["M3-2","M3-4"],["PRPS"],0,"Service Manager"),
 # Module 4
 d("POL-SCO-01","Policy","Specialised Support Coordination Policy","Support coordination","M4","P2",["M4-1","M4-2","M4-3"],["PRPS"]),
 d("PRO-SCO-01","Procedure","Support Coordination Conflict of Interest Procedure","Support coordination","M4","P2",["M4-3"],["PRPS"],0,"Service Manager"),
 # Module 5
 d("POL-SDA-01","Policy","Specialist Disability Accommodation Policy","SDA","M5","P2",["M5-1","M5-2","M5-3","M5-4","M5-5"],["SDAC","PRPS"]),
 d("PRO-SDA-01","Procedure","SDA Dwelling Enrolment and Compliance Procedure","SDA","M5","P2",["M5-4","M5-5"],["SDAC"],0,"Service Manager"),
 d("REG-SDA-01","Register","SDA Dwelling Register","SDA","M5","P2",["M5-4"],["SDAC"],0,"Service Manager",1),
 # SIL module
 d("PRO-SIL-01","Procedure","Supported Decision-Making Procedure","SIL","SIL","P1",["SIL-1"],["PRPS26"],1,"House Supervisor"),
 d("FRM-SIL-01","Form","Decision Support Record","SIL","SIL","P1",["SIL-1"],["PRPS26"],1,"House Supervisor",1),
 d("POL-SIL-02","Policy","Safeguarding in the Home Policy (SIL)","SIL","SIL","P1",["SIL-2"],["PRPS26","IMRI"],1),
 d("PRO-SIL-02","Procedure","Home Risk, Dignity of Risk and Safeguarding Procedure","SIL","SIL","P1",["SIL-2"],["PRPS26"],1,"House Supervisor"),
 d("POL-SIL-03","Policy","SIL Practice Governance Policy","SIL","SIL","P1",["SIL-3"],["PRPS26","QI26"]),
 d("PRO-SIL-03","Procedure","House Governance, Handover and Worker Training Procedure","SIL","SIL","P1",["SIL-3"],["PRPS26"],0,"House Supervisor"),
 d("PRO-SIL-04","Procedure","Tenancy and Support Agreements Procedure (SIL)","SIL","SIL","P1",["SIL-4"],["PRPS26"],1,"Service Manager"),
 d("AGR-SIL-01","Agreement","SIL Tenancy and Support Agreement Templates","SIL","SIL","P1",["SIL-4"],["PRPS26"],1,"Service Manager",1),
 d("REG-SIL-01","Register","SIL Homes Register","SIL","SIL","P2",["SIL-3"],["PRPS26"],0,"Service Manager",1),
 # Adjacent
 d("POL-WHS-01","Policy","Work Health and Safety Policy","Environment","All","P2",["CORE-4.1"],["WHSL"]),
]

# ---------------- STRUCTURAL DOCUMENT EDGES ----------------
E = []
def e(f,t,to,note=""): E.append((f,t,to,note))
OP = [("PRO-GOV-01","POL-GOV-01"),("PRO-GOV-02","POL-GOV-01"),("PRO-RSK-01","POL-RSK-01"),("PRO-QMS-01","POL-QMS-01"),
 ("PRO-QMS-02","POL-QMS-01"),("PRO-QMS-03","POL-QMS-01"),("PRO-INF-01","POL-INF-01"),
 ("PRO-INF-02","POL-INF-01"),("PRO-FBK-01","POL-FBK-01"),("WIN-FBK-01","PRO-FBK-01"),
 ("PRO-INC-01","POL-INC-01"),("PRO-INC-02","POL-INC-01"),("WIN-INC-01","PRO-INC-01"),
 ("PRO-HRM-01","POL-HRM-01"),("PRO-HRM-02","POL-HRM-01"),("PRO-HRM-03","POL-HRM-01"),
 ("PRO-COS-01","POL-COS-01"),("PLN-COS-01","POL-COS-01"),("PRO-EDM-01","POL-EDM-01"),
 ("PLN-EDM-01","POL-EDM-01"),("PRO-SGD-01","POL-SGD-01"),("PRO-SUP-01","POL-SUP-01"),
 ("PRO-SUP-02","POL-SUP-02"),("PRO-SUP-03","POL-SUP-03"),("PRO-SUP-04","POL-SUP-04"),
 ("PRO-ENV-01","POL-ENV-01"),("PRO-MNY-01","POL-MNY-01"),("PRO-MED-01","POL-MED-01"),
 ("WIN-MED-01","PRO-MED-01"),("PRO-MTM-01","POL-MTM-01"),("PRO-WST-01","POL-WST-01"),
 ("PRO-HID-01","POL-HID-01"),("PRO-HID-02","POL-HID-01"),("PRO-HID-03","POL-HID-01"),
 ("PRO-HID-04","POL-HID-01"),("PRO-HID-05","POL-HID-01"),("PRO-HID-06","POL-HID-01"),
 ("PRO-HID-07","POL-HID-01"),("PRO-HID-08","POL-HID-01"),("PRO-BSP-01","POL-BSP-01"),
 ("PRO-BSP-02","POL-BSP-01"),("PRO-BSP-03","POL-BSP-01"),("PRO-ECS-01","POL-ECS-01"),
 ("PRO-SCO-01","POL-SCO-01"),("PRO-SDA-01","POL-SDA-01"),("PRO-SIL-01","POL-RGT-03"),
 ("PRO-SIL-02","POL-SIL-02"),("PRO-SIL-03","POL-SIL-03"),("PRO-SIL-04","POL-SUP-03")]
for f,t in OP: e(f,"OPERATIONALISES",t)
USES = [("PRO-GOV-01","FRM-GOV-01"),("PRO-RSK-01","FRM-RSK-01"),("PRO-QMS-01","FRM-QMS-01"),
 ("PRO-INF-01","FRM-INF-01"),("PRO-FBK-01","FRM-FBK-01"),("PRO-INC-01","FRM-INC-01"),
 ("PRO-HRM-01","FRM-HRM-01"),("PRO-HRM-01","FRM-HRM-02"),("PRO-SUP-01","FRM-SUP-01"),
 ("PRO-SUP-02","FRM-SUP-02"),("PRO-SUP-04","FRM-SUP-04"),("PRO-ENV-01","FRM-ENV-01"),
 ("PRO-MNY-01","FRM-MNY-01"),("PRO-MED-01","FRM-MED-01"),("PRO-MTM-01","FRM-MTM-01"),
 ("PRO-HID-01","FRM-HID-01"),("PRO-BSP-02","FRM-BSP-01"),("PRO-SIL-01","FRM-SIL-01"),
 ("PRO-SUP-03","AGR-SUP-01"),("PRO-SIL-04","AGR-SIL-01"),("PRO-RGT? ","")]
USES = [u for u in USES if u[1]]
for f,t in USES: e(f,"USES",t)
REC = [("FRM-GOV-01","REG-GOV-02"),("FRM-RSK-01","REG-RSK-01"),("FRM-QMS-01","REG-QMS-01"),
 ("FRM-FBK-01","REG-FBK-01"),("FRM-INC-01","REG-INC-01"),("FRM-HRM-01","REG-HRM-01"),
 ("FRM-SUP-02","REG-SUP-01"),("AGR-SUP-01","REG-SUP-01"),("FRM-ENV-01","REG-ENV-01"),
 ("FRM-MNY-01","REG-MNY-01"),("FRM-BSP-01","REG-BSP-01"),("AGR-SIL-01","REG-SIL-01")]
for f,t in REC: e(f,"RECORDS_TO",t)
REL = [("POL-GOV-03","PRO-FBK-01","Disclosures may arrive via the complaints channel; both must protect the discloser"),
 ("PRO-INF-02","PRO-INC-01","A privacy breach is also managed as an incident"),
 ("PRO-SGD-01","PRO-INC-02","Abuse or neglect triggers reportable-incident notification"),
 ("PRO-BSP-02","PRO-INC-02","Unauthorised restrictive practice is a reportable incident"),
 ("PRO-MED-01","PRO-INC-01","Medication errors are recorded and managed as incidents"),
 ("POL-RGT-03","PRO-SIL-01","Supported decision-making applies org-wide and in SIL homes"),
 ("POL-WHS-01","POL-ENV-01","WHS duties and safe-environment obligations overlap"),
 ("PRO-MTM-01","PRO-HID-03","Mealtime management escalates to dysphagia procedure where assessed")]
for f,to,n in REL: e(f,"RELATES_TO",to,n)

# ---------------- BUILD ----------------
LEGK = {l[0]: l for l in LEG}; STDK = {s[0]: s for s in STD}; DOCK = {x["id"]: x for x in DOCS}
# implements edges + statutory-reference edges
for x in DOCS:
    for s in x["imp"]: e(x["id"],"IMPLEMENTS",s)
    for l in x["leg"]: e(x["id"],"RELATES_TO",l,"Statutory reference (proposed REFERENCES type - taxonomy v0.4 candidate)")
# legislative / framework layer edges
for m in MODULES: e(m[0],"PART_OF","NDIS-PS","Module of the NDIS Practice Standards")
e("NDIS-PS","PUBLISHED_IN","PRPS","Practice Standards are set out in the Rules (delegated legislation)")
for s in STD: e(s[0],"PART_OF",s[1])
for k in ["PRPS","IMRI","CMR","RPBS","WSR","COC","SDAC","PD","QI","QI26","PRPS26","AQA","AQAR25","NCE","BSPA"]:
    e(k,"MADE_UNDER","ACT")
e("PRPS26","RELATES_TO","PRPS","Amending instrument (proposed AMENDS type - taxonomy v0.4 candidate)")
e("QI26","RELATES_TO","QI","Amending instrument (proposed AMENDS type - taxonomy v0.4 candidate)")
e("IS26","RELATES_TO","ACT","Amending Act (proposed AMENDS type - taxonomy v0.4 candidate)")
e("SIL","RELATES_TO","PRPS26","SIL module introduced by the 2026 Amendment Rules")

# validation
ids = set(DOCK)|set(STDK)|set(LEGK)|{m[0] for m in MODULES}|{"NDIS-PS"}
bad = [(f,t,to) for f,t,to,_ in E if f not in ids or to not in ids]
assert not bad, f"Dangling edges: {bad}"

outg = {}; inc = {}
for f,t,to,n in E:
    if f in DOCK and (to in DOCK or to in STDK):
        outg.setdefault(f,[]).append(f"{t} {to}")
    if to in DOCK and f in DOCK:
        inc.setdefault(to,[]).append(f"{f} {t} this")

# ---------------- XLSX ----------------
wb = Workbook(); AR = Font(name="Arial", size=10)
H = Font(name="Arial", size=10, bold=True, color="FFFFFF")
FILL = PatternFill("solid", start_color="1F3864")
FILL2 = PatternFill("solid", start_color="DDEBF7")
WRAP = Alignment(wrap_text=True, vertical="top")
thin = Border(*[Side(style="thin", color="B0B0B0")]*4)

def sheet(ws, headers, widths):
    for c,(h,w) in enumerate(zip(headers,widths),1):
        cell = ws.cell(1,c,h); cell.font=H; cell.fill=FILL; cell.alignment=WRAP
        ws.column_dimensions[get_column_letter(c)].width=w
    ws.freeze_panes="A2"

rm = wb.active; rm.title="READ ME"
rm.column_dimensions["A"].width=30; rm.column_dimensions["B"].width=100
rows = [
 ("RISE - NDIS PROVIDER DOCUMENT REGISTER",""),
 ("Purpose","Master map of a best-practice NDIS provider document suite: every policy, procedure, work instruction, form, register, plan and agreement, with typed links to the Practice Standards and legislation. This workbook is itself REG-GOV-03 in the suite and doubles as the human-readable node/edge store for the Rise compliance graph."),
 ("Generated",TODAY+" | Suite design v0.1 | Companion machine file: rise-nodes-and-edges.json | Edge types per rise-relationship-taxonomy.md v0.3"),
 ("",""),
 ("HOW TO READ THE TYPED LINKS (plain English)",""),
 ("IMPLEMENTS","This document exists to meet that Practice Standard."),
 ("OPERATIONALISES","This procedure/work instruction puts that policy/procedure into day-to-day practice."),
 ("USES","This procedure is carried out using that form."),
 ("RECORDS_TO","Completed copies of this form/agreement are logged in that register."),
 ("PART_OF / PUBLISHED_IN / MADE_UNDER","Standard sits in a module; the Standards are set out in the 2018 Rules; the Rules are made under the NDIS Act 2013."),
 ("RELATES_TO","Governed escape hatch (always carries a note). Statutory references and amendments use it pending proposed edge types REFERENCES and AMENDS (taxonomy v0.4 candidates)."),
 ("",""),
 ("SHEETS",""),
 ("Documents","The full suite ({} documents). Filters on. Priority: P1 = audit-critical first wave, P2 = core completion, P3 = supporting."),
 ("Standards","Every Practice Standard node with its module, division and verification status."),
 ("Legislation","Instrument nodes (Acts, Rules, Guidelines) and their role in the graph."),
 ("Edges","Every link as From / Type / To / Note - machine-readable mirror of the JSON."),
 ("",""),
 ("COUNTS",""),
 ("Policies",'=COUNTIF(Documents!B:B,"Policy")'),
 ("Procedures",'=COUNTIF(Documents!B:B,"Procedure")'),
 ("Work instructions",'=COUNTIF(Documents!B:B,"Work Instruction")'),
 ("Forms",'=COUNTIF(Documents!B:B,"Form")+COUNTIF(Documents!B:B,"Agreement")'),
 ("Registers",'=COUNTIF(Documents!B:B,"Register")'),
 ("Plans / handbooks",'=COUNTIF(Documents!B:B,"Plan")+COUNTIF(Documents!B:B,"Handbook")'),
 ("Total documents","=COUNTA(Documents!A:A)-1"),
 ("Total edges","=COUNTA(Edges!A:A)-1"),
 ("",""),
 ("VERIFICATION LEGEND (Standards & Legislation sheets)",""),
 ("V","Verified against the NDIS Commission website or instrument listing, 5 Jul 2026."),
 ("C","Established knowledge corroborated by secondary sources; low risk."),
 ("T","Training knowledge - confirm against the current compilation on the Federal Register of Legislation before relying on it externally."),
 ("D","Draft-source. As of 6 Jul 2026 none remain at standard level: SIL names verified via QI Guidelines compilation C03. PRPS clause pinpoints for SIL still await the next PRPS compilation and stay TBC in the Citation column."),
 ("",""),
 ("EPISTEMIC NOTES","The suite composition, IDs, priorities, owners and review cycles are design recommendations, not regulatory requirements. Which modules apply to a given provider depends on their registration groups. Core Module standard names reflect the 2021 v4 structure; the 2026 Amendment Rules and any Practice Standards Review outcomes should be checked against this register before an audit. Schedule/clause citations on the Standards sheet come from the design-stage compilation review; spot-check clause numbers and PRPS body-section references against the live Federal Register compilation before external publication."),
]
rows[13] = ("Documents", rows[13][1].format(len(DOCS)))
for r,(a,b) in enumerate(rows,1):
    ca=rm.cell(r,1,a); cb=rm.cell(r,2,b); ca.font=Font(name="Arial",size=10,bold=True); cb.font=AR
    ca.alignment=WRAP; cb.alignment=WRAP
rm.cell(1,1).font=Font(name="Arial", size=14, bold=True, color="1F3864")

ws = wb.create_sheet("Documents")
hdr=["Doc ID","Type","Title","Cluster","Applies to","Priority","Implements (standards)","Key legislation","Links out (typed)","Linked from (typed)","Easy Read","Owner (role)","Review (yrs)","Status"]
sheet(ws,hdr,[12,13,44,15,16,8,22,30,34,34,9,16,9,10])
for x in DOCS:
    leg = "; ".join(LEGK[k][1].replace("National Disability Insurance Scheme","NDIS") for k in x["leg"])
    row=[x["id"],x["type"],x["title"],x["cluster"],x["applies"],x["priority"],
         "; ".join(x["imp"]),leg,"\n".join(outg.get(x["id"],[])),"\n".join(inc.get(x["id"],[])),
         "Yes" if x["ez"] else "",x["own"],x["rv"],"Planned"]
    ws.append(row)
for r in ws.iter_rows(min_row=2):
    for c in r: c.font=AR; c.alignment=WRAP; c.border=thin
    if r[0].row%2==0:
        for c in r: c.fill=FILL2
ws.auto_filter.ref = f"A1:N{ws.max_row}"

ws = wb.create_sheet("Standards")
sheet(ws,["Std ID","Module","Division","Standard","Citation - PRPS Rules 2018 as amended","Quality indicators - QI Guidelines C03","Verification"],[10,10,30,42,32,16,12])
for s in STD:
    ws.append([s[0],s[1],s[2],s[3],CIT.get(s[0],""),QIREF.get(s[0],""),s[4]])
for r in ws.iter_rows(min_row=2):
    for c in r: c.font=AR; c.alignment=WRAP; c.border=thin
ws.auto_filter.ref=f"A1:F{ws.max_row}"

ws = wb.create_sheet("Legislation")
sheet(ws,["Key","Instrument","Kind","Notes","Verification","Role in graph"],[8,52,20,42,12,30])
for l in LEG: ws.append(list(l[:2])+[l[2],l[3],l[4],l[5]])
for r in ws.iter_rows(min_row=2):
    for c in r: c.font=AR; c.alignment=WRAP; c.border=thin

ws = wb.create_sheet("Edges")
sheet(ws,["From","Edge type","To","Note"],[16,18,16,60])
for f,t,to,n in E: ws.append([f,t,to,n])
for r in ws.iter_rows(min_row=2):
    for c in r: c.font=AR; c.alignment=WRAP; c.border=thin
ws.auto_filter.ref=f"A1:D{ws.max_row}"

wb.save("/home/claude/rise-document-register.xlsx")

# ---------------- JSON ----------------
nodes=[]
for l in LEG: nodes.append(dict(id=l[0],node_type="Instrument",kind=l[2],title=l[1],layer="shared",verification=l[4],note=l[3]))
nodes.append(dict(id="NDIS-PS",node_type="Framework",title="NDIS Practice Standards",layer="shared",verification="V"))
for m in MODULES: nodes.append(dict(id=m[0],node_type="Module",title=m[1],applies_to=m[2],layer="shared",verification=m[3]))
for s in STD: nodes.append(dict(id=s[0],node_type="Standard",title=s[3],module=s[1],division=s[2],citation=CIT.get(s[0],""),qi_ref=QIREF.get(s[0],""),layer="shared",verification=s[4]))
for x in DOCS: nodes.append(dict(id=x["id"],node_type="Document",doc_type=x["type"],title=x["title"],cluster=x["cluster"],applies_to=x["applies"],priority=x["priority"],easy_read=bool(x["ez"]),owner_role=x["own"],review_years=x["rv"],status="Planned",layer="tenant"))
edges=[dict(source=f,type=t,target=to,**({"note":n} if n else {})) for f,t,to,n in E]
meta=dict(version="0.1",generated=TODAY,edge_taxonomy="rise-relationship-taxonomy.md v0.3",
 assertion_source="human-design",verification_legend=dict(V="verified vs regulator/primary source Jul 2026",C="corroborated training knowledge",T="training knowledge - confirm vs current compilation",D="draft instrument - confirm final wording"))
json.dump(dict(meta=meta,nodes=nodes,edges=edges),open("/home/claude/rise-nodes-and-edges.json","w"),indent=1)
print("docs",len(DOCS),"standards",len(STD),"instruments",len(LEG),"edges",len(E),"nodes",len(nodes))
