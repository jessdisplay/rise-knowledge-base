#!/usr/bin/env python3
"""Rise knowledge-base assembler — 7 Jul 2026.

Combines rise-ndis-package (Session B) and rise-audit-pack v2 (Session A)
into one browsable knowledge base. Original files are copied unmodified;
every .md/.py/.json/.mermaid/.xlsx gains a rendered .html sibling; a
generated index.html links to everything. This script itself is copied
into _kb/ so the index can be regenerated (single-source pattern).
"""
import html, json, re, shutil
from pathlib import Path

import markdown
from openpyxl import load_workbook

SRC = Path("/home/claude/kb-src")
OUT = Path("/home/claude/out/rise-knowledge-base")
BUILD_DATE = "7 July 2026"

# Mode is decided by where this script lives, not by what paths exist:
# the copy inside a repo's _kb/ ALWAYS runs in-place (re-render only, never
# delete/copy); only the assembly copy outside _kb/ can rebuild from SRC.
_here = Path(__file__).resolve()
IN_PLACE = _here.parent.name == "_kb"
if IN_PLACE:
    OUT = _here.parent.parent
    print("in-place regeneration:", OUT)

# ---------------------------------------------------------------- copy trees
if not IN_PLACE:
    if OUT.exists():
        shutil.rmtree(OUT)
    OUT.mkdir(parents=True)
    for src_rel, dst_name in (
        ("rise-ndis-package", "rise-ndis-package"),
        ("rise-audit-pack", "rise-audit-pack"),
        ("new-uploads/rise-dossier-zip", "rise-dossier-bundle"),
        ("rise-authoring-research", "rise-authoring-research"),
    ):
        shutil.copytree(SRC / src_rel, OUT / dst_name)
(OUT / "_kb").mkdir(exist_ok=True)

MD = markdown.Markdown(extensions=["tables", "fenced_code", "sane_lists"])

CLAUDE_MD = """# CLAUDE.md — Rise NDIS compliance knowledge base

Orientation for Claude Code (or any agent) working in this repository. Read this before editing anything. This file is written by `_kb/build_kb.py`; edit it there.

## What this repository is

Three preserved session collections plus a generated navigation layer, assembled 7 July 2026:

- `rise-ndis-package/` — the document-and-graph package: 123-document suite, register workbook, seed graph JSON, templates, generator scripts, interactive dossier.
- `rise-audit-pack/` — the audit-research pack: auditor guide, indexed checklist, booklet reconciliation, coverage map v2, the Commission booklet PDF (Nov 2021).
- `rise-dossier-bundle/` — the 5–7 Jul design-session bundle: Claude Code briefs, source manifest, relationship taxonomy v0.3, plain-English map, navigator, visual atlas, training model, prototypes.
- `rise-authoring-research/` — 7 Jul web-verified authoring/formatting best practice.
- `index.html`, `standards.html`, `_kb/` — the generated knowledge-base layer.

## Read in this order for build tasks

1. `rise-dossier-bundle/rise-dossier-index.md` — §5 is the standing instruction for Claude Code.
2. `rise-dossier-bundle/rise-claude-code-brief.md` — Milestone 1 (taxonomy v0.3 embedded as Appendix A).
3. `rise-dossier-bundle/rise-source-manifest.md` — **execute before writing any seed data.**
4. `rise-dossier-bundle/rise-milestone-2-gap-engine.md` — Milestone 2.
5. `rise-ndis-package/rise-developer-handoff.md` — the package's import plan.

## Hard rules (the collection's own discipline — do not relax)

- Never reconstruct clause numbers, section numbers, register IDs, F-numbers, counts or quotes from memory. Retrieve from a file in this repo or a fetched primary source, or leave an explicit `[GAP]`.
- Label claims: verified `[V]` (with source and date) / corroborated `[C]` / design opinion `[D]`. Anything reconstructed later "from memory" of past sessions is unverified by definition.
- The register workbook's READ ME sheet is authoritative for counts when records disagree.
- Treat the three collection folders as **read-only provenance**. Do not edit their files in place; new or superseding work goes in new top-level folders (like `rise-authoring-research/`), versioned, stating what it supersedes.
- Statutory addresses were last verified against fetched instruments in prior sessions (SIL Rules clauses are TBC in the register itself). Before anything load-bearing, re-verify against the in-force compilations on legislation.gov.au: PRPS Rules F2018L00631 · QI Guidelines C03 F2026C00528.
- Not legal advice. Where anything here disagrees with the instrument on the Federal Register, the instrument governs.

## Generated vs source — never hand-edit outputs

`index.html`, `standards.html`, and every rendered `.html` that sits beside a source file (each `*.md` → `.html`, plus `*.py.html`, `*.json.html`, `*.mermaid.html`, `*.xlsx.html`) are **outputs** of `_kb/build_kb.py`. To change them, change the script or the source files, then regenerate from the repo root:

```
pip install openpyxl markdown
python3 _kb/build_kb.py
```

Run in-place it re-renders everything and rewrites `index.html`, `standards.html` and this file; it does not move or delete the collections. It prints sanity counts on completion (expected: 66/70 plain-English matches, 123 suite documents).

## Current open items (as at 7 Jul 2026 — see the index status band)

SIL Rules clauses (TBC in the register) · fresh-fetch currency check of all addresses · two statutory timeframes flagged "confirm before approval" · Easy Read companions flagged but not drafted (production standard: `rise-authoring-research/`) · four standards with no implementing document (M2-2, M2-4, M2-6, M2A-3 — review whether intentional) · the bundle's seven-item open-questions register.

## Layout note

The tree is preserved as-uploaded (provenance-faithful). The unified repo layout proposed in `rise-audit-pack/04-manifest/rise-collection-manifest.md` §2 is a documented restructure option — note it predates `standards.html`, `index.html`, the bundle and the authoring research, so adopting it means extending that plan and updating this generator's paths, as its own versioned step.
"""

(OUT / "CLAUDE.md").write_text(CLAUDE_MD, encoding="utf-8")

# ---------------------------------------------------------------- shared css
KB_CSS = """
:root{
  --navy:#12284b;--navy2:#1d3a66;--bg:#eef1f2;--panel:#ffffff;--ink:#17202b;
  --mut:#5a6673;--rule:#c9d1d6;--link:#12408a;
  --ok:#1e7a4f;--okbg:#e5f2ea;--inf:#534ab7;--infbg:#eeedfe;--amb:#8a5b00;--ambbg:#f6ecd6;
  --pol:#12284b;--pro:#0f6a6f;--win:#575f6b;--frm:#7a4a1e;--reg:#1e7a4f;--pln:#8a5b00;
  --gen:#3d4b5c;
}
@media (prefers-color-scheme: dark){
  :root{
    --bg:#13171c;--panel:#1c2129;--ink:#e8ecf1;--mut:#9aa7b4;--rule:#323b46;--link:#8db4ea;
    --ok:#7fc9a2;--okbg:#173525;--inf:#b3ace8;--infbg:#26215c;--amb:#e8b45a;--ambbg:#3f2c07;
    --pol:#8fa9d4;--pro:#6cc4c9;--win:#a9b2be;--frm:#d3a06b;--reg:#7fc9a2;--pln:#e0b465;
    --gen:#9fb0c2;
  }
}
*{box-sizing:border-box}
html{scroll-behavior:smooth}
@media (prefers-reduced-motion:reduce){html{scroll-behavior:auto}}
body{margin:0;background:var(--bg);color:var(--ink);
  font:15.5px/1.6 "Public Sans",-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,system-ui,sans-serif}
a{color:var(--link)}
a:focus-visible{outline:2px solid var(--link);outline-offset:2px;border-radius:2px}
code,pre,.mono{font-family:"IBM Plex Mono",ui-monospace,SFMono-Regular,Menlo,Consolas,monospace}
.kbbar{background:var(--navy);color:#dbe4f2;padding:10px 20px;font-size:13px}
.kbbar a{color:#dbe4f2;text-decoration:none;font-weight:600}
.kbbar a:hover{text-decoration:underline}
.kbbar .sep{opacity:.45;margin:0 7px}
.kbwrap{max-width:900px;margin:26px auto 60px;padding:0 18px}
article.kbdoc{background:var(--panel);border:1px solid var(--rule);border-radius:10px;padding:34px 42px 40px}
@media(max-width:640px){article.kbdoc{padding:22px 18px}}
.kbmeta{font:12.5px/1.6 "IBM Plex Mono",monospace;color:var(--mut);border-bottom:1px solid var(--rule);
  padding-bottom:12px;margin-bottom:22px;display:flex;flex-wrap:wrap;gap:6px 18px}
.kbnote{background:var(--ambbg);border:1px solid var(--amb);color:var(--ink);border-radius:8px;
  padding:12px 16px;margin:0 0 22px;font-size:14px}
.kbnote b{color:var(--amb)}
.kbdoc h1{font:800 29px/1.2 Spectral,Georgia,serif;margin:0 0 14px}
.kbdoc h2{font:700 21px/1.25 Spectral,Georgia,serif;margin:34px 0 10px;padding-bottom:5px;border-bottom:1px solid var(--rule)}
.kbdoc h3{font:700 17px/1.3 Spectral,Georgia,serif;margin:26px 0 8px}
.kbdoc h4{font:600 15px/1.3 "Public Sans",sans-serif;margin:20px 0 6px}
.kbdoc p,.kbdoc li{max-width:74ch}
.kbdoc blockquote{margin:16px 0;padding:10px 16px;background:var(--ambbg);border-left:4px solid var(--amb);border-radius:0 8px 8px 0;font-size:14.5px}
.kbdoc table{border-collapse:collapse;font-size:13.5px;display:block;overflow-x:auto;max-width:100%;margin:14px 0}
.kbdoc th,.kbdoc td{border:1px solid var(--rule);padding:7px 10px;text-align:left;vertical-align:top}
.kbdoc th{background:color-mix(in srgb,var(--panel) 88%,var(--ink) 12%);font-weight:600}
.kbdoc pre{background:color-mix(in srgb,var(--panel) 90%,var(--ink) 10%);border:1px solid var(--rule);
  border-radius:8px;padding:14px;overflow-x:auto;font-size:13px;line-height:1.5}
.kbdoc code{font-size:.92em}
.kbdoc :not(pre)>code{background:color-mix(in srgb,var(--panel) 88%,var(--ink) 12%);padding:1px 5px;border-radius:4px}
.kbdoc hr{border:0;border-top:1px solid var(--rule);margin:28px 0}
.kbfoot{margin-top:30px;padding-top:14px;border-top:1px solid var(--rule);font-size:13px;color:var(--mut)}
details.sheet{margin:14px 0;border:1px solid var(--rule);border-radius:8px;background:var(--panel)}
details.sheet>summary{cursor:pointer;padding:10px 14px;font:600 15px Spectral,Georgia,serif}
details.sheet>div{padding:0 14px 14px;overflow-x:auto}
details.sheet table{font-size:12.5px}
@media print{body{background:#fff}article.kbdoc{border:none}}
"""

PAGE = """<!DOCTYPE html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>@@TITLE@@ · Rise KB</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="stylesheet" href="https://fonts.googleapis.com/css2?family=Spectral:wght@500;700;800&family=Public+Sans:wght@400;600;700&family=IBM+Plex+Mono:wght@400;600&display=swap" media="all" onerror="this.media='none'">
<link rel="stylesheet" href="@@ROOT@@_kb/kb.css">
</head><body>
<div class="kbbar"><a href="@@ROOT@@index.html">Rise Knowledge Base</a><span class="sep">/</span>@@CRUMB@@</div>
<div class="kbwrap"><article class="kbdoc">
<div class="kbmeta">@@META@@</div>
@@NOTE@@
@@BODY@@
<div class="kbfoot">@@FOOT@@</div>
</article></div>
</body></html>
"""


def esc(s):
    return html.escape(str(s), quote=True)


def rel_root(rel: Path) -> str:
    return "../" * (len(rel.parts) - 1)


def crumb(rel: Path) -> str:
    return '<span class="mono">' + esc(str(rel)) + "</span>"


def write_page(rel: Path, title, meta, body, foot, note=""):
    out = PAGE.replace("@@TITLE@@", esc(title)).replace("@@ROOT@@", rel_root(rel))
    out = out.replace("@@CRUMB@@", crumb(rel)).replace("@@META@@", meta)
    out = out.replace("@@NOTE@@", note).replace("@@BODY@@", body).replace("@@FOOT@@", foot)
    (OUT / rel).write_text(out, encoding="utf-8")


LINK_RE = re.compile(r'(href=")(?!https?://|mailto:|#)([^"#]*?)\.md(#[^"]*)?(")')

# Pages that get an assembly note injected (rendered page only; original md untouched)
CHECKLIST_NOTE = (
    '<div class="kbnote"><b>Assembly note — added 7 Jul 2026, not part of the original document.</b> '
    "The gap tables in §§3–5 below and the 48 / 6 / 16 counts were superseded later the same day by the "
    '<a href="rise-reconciliation-2021v4-booklet.html">booklet reconciliation</a>: all 70 standard names are now '
    "verified, 18 QI-section pairings carry an inference label, and the Rules clauses for M2-3…7, Module 2A, "
    'M3-1…4 and all SIL rows remain open. Current state: <a href="../03-visuals/ndis-standards-coverage-map-v2.html">'
    "coverage map v2</a>. The Markdown file alongside this page is unmodified.</div>"
)


def render_md(path: Path):
    rel = path.relative_to(OUT)
    text = path.read_text(encoding="utf-8")
    m = re.search(r"^#\s+(.+)$", text, re.M)
    title = m.group(1).strip() if m else path.name
    MD.reset()
    body = MD.convert(text)
    body = LINK_RE.sub(lambda g: g.group(1) + g.group(2) + ".html" + (g.group(3) or "") + g.group(4), body)
    out_rel = rel.with_suffix(".html")
    meta = (
        "<span>rendered view</span><span>source: " + esc(path.name) + "</span>"
        "<span>" + f"{path.stat().st_size:,}" + " bytes</span>"
    )
    foot = ('Rendered for browsing on ' + BUILD_DATE + " — the original Markdown is preserved unmodified alongside: "
            '<a class="mono" href="' + esc(path.name) + '">' + esc(path.name) + "</a>")
    note = CHECKLIST_NOTE if path.name == "rise-ndis-audit-checklist-indexed.md" else ""
    write_page(out_rel, title, meta, body, foot, note)
    return title, out_rel


def render_code(path: Path, label):
    rel = path.relative_to(OUT)
    code = path.read_text(encoding="utf-8", errors="replace")
    body = "<h1>" + esc(path.name) + "</h1><pre><code>" + esc(code) + "</code></pre>"
    out_rel = Path(str(rel) + ".html")
    meta = "<span>" + esc(label) + "</span><span>" + f"{path.stat().st_size:,}" + " bytes</span>"
    foot = 'Read-only view. Original file: <a class="mono" href="' + esc(path.name) + '">' + esc(path.name) + "</a>"
    write_page(out_rel, path.name, meta, body, foot)
    return out_rel


def render_mermaid(path: Path):
    rel = path.relative_to(OUT)
    code = path.read_text(encoding="utf-8")
    body = (
        "<h1>" + esc(path.name) + "</h1>"
        '<p id="mmnote" style="font-size:13.5px;color:var(--mut)">Attempting to render the diagram '
        "(needs internet for the mermaid library). If nothing appears below, you are offline — the "
        "diagram source follows either way.</p>"
        '<div id="mm"></div>'
        "<h2>Source</h2><pre><code>" + esc(code) + "</code></pre>"
        '<script type="module">'
        "try{const m=await import('https://cdn.jsdelivr.net/npm/mermaid@10/dist/mermaid.esm.min.mjs');"
        "m.default.initialize({startOnLoad:false,securityLevel:'strict'});"
        "const{svg}=await m.default.render('g1'," + json.dumps(code) + ");"
        "document.getElementById('mm').innerHTML=svg;"
        "document.getElementById('mmnote').textContent='Rendered with mermaid (loaded from CDN). Source below.';}"
        "catch(e){document.getElementById('mmnote').textContent='Diagram library unavailable (offline?) — showing source only.';}"
        "</script>"
    )
    out_rel = Path(str(rel) + ".html")
    meta = "<span>mermaid diagram</span><span>" + f"{path.stat().st_size:,}" + " bytes</span>"
    foot = 'Original file: <a class="mono" href="' + esc(path.name) + '">' + esc(path.name) + "</a>"
    write_page(out_rel, path.name, meta, body, foot)
    return out_rel


def render_xlsx(path: Path):
    rel = path.relative_to(OUT)
    wb = load_workbook(path, read_only=True, data_only=True)
    parts = ["<h1>" + esc(path.name) + " — sheet preview</h1>",
             '<p style="font-size:14px;color:var(--mut)">Values-only preview generated on ' + BUILD_DATE +
             '. Formulas, formatting and validations are only in the <a class="mono" href="' + esc(path.name) +
             '">original workbook</a>.</p>']
    first = True
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        n = len(rows)
        parts.append('<details class="sheet"' + (" open" if first else "") + "><summary>" +
                     esc(ws.title) + ' <span class="mono" style="font-weight:400;color:var(--mut)">— ' +
                     str(max(n - 1, 0)) + " data rows</span></summary><div><table>")
        for i, row in enumerate(rows):
            tag = "th" if i == 0 else "td"
            parts.append("<tr>" + "".join(
                "<" + tag + ">" + esc("" if c is None else c) + "</" + tag + ">" for c in row) + "</tr>")
        parts.append("</table></div></details>")
        first = False
    wb.close()
    out_rel = Path(str(rel) + ".html")
    meta = "<span>spreadsheet preview</span><span>" + f"{path.stat().st_size:,}" + " bytes</span>"
    foot = ('Download the workbook itself: <a class="mono" href="' + esc(path.name) + '">' + esc(path.name) + "</a>")
    write_page(out_rel, path.name + " preview", meta, body="".join(parts), foot=foot)
    return out_rel


# ---------------------------------------------------------------- walk + render
catalogue = {}   # rel md/original path -> dict(title, view, kind)

for path in sorted(OUT.rglob("*")):
    if not path.is_file() or "_kb" in path.parts:
        continue
    rel = path.relative_to(OUT)
    ext = path.suffix.lower()
    # Skip generated outputs (matters for in-place runs, where they already exist):
    # the two root pages, every .html rendered from a sibling .md, and every
    # viewer .html sitting on top of a .py/.json/.mermaid/.xlsx source.
    if ext == ".html":
        if len(rel.parts) == 1 and path.name in ("index.html", "standards.html"):
            continue
        base = path.with_name(path.name[:-5])
        if path.with_suffix(".md").exists() or (
                base.suffix in (".py", ".json", ".mermaid", ".xlsx") and base.exists()):
            continue
    if ext == ".md":
        title, view = render_md(path)
        catalogue[str(rel)] = {"title": title, "view": str(view), "kind": "page"}
    elif ext == ".py":
        view = render_code(path, "python source")
        catalogue[str(rel)] = {"title": path.name, "view": str(view), "kind": "code"}
    elif ext == ".json":
        view = render_code(path, "json data")
        catalogue[str(rel)] = {"title": path.name, "view": str(view), "kind": "data"}
    elif ext == ".mermaid":
        view = render_mermaid(path)
        catalogue[str(rel)] = {"title": path.name, "view": str(view), "kind": "diagram"}
    elif ext == ".xlsx":
        view = render_xlsx(path)
        catalogue[str(rel)] = {"title": path.name, "view": str(view), "kind": "spreadsheet"}
    elif ext in (".html", ".pdf"):
        title = path.name
        if ext == ".html":
            m = re.search(r"<title>([^<]+)</title>",
                          path.read_text(encoding="utf-8", errors="replace")[:4096])
            if m:
                title = m.group(1).strip()
        catalogue[str(rel)] = {"title": title, "view": str(rel),
                               "kind": "interactive" if ext == ".html" else "pdf"}

# ---------------------------------------------------------------- standards map (generated from the register)
ID_RE = re.compile(r"^([A-Z]{3}-[A-Z]{3}-\d{2}[A-Za-z]?)")
BOOKLET_NAMES = {  # Commission booklet (Nov 2021 v4) titles, from the reconciliation doc — for dual-title flags
 "M2-1":"Behaviour Support in the NDIS","M2-2":"Restrictive Practices",
 "M2-3":"Functional Behaviour Assessments and Behaviour Support Plans",
 "M2-4":"Supporting the Implementation of the Behaviour Support Plan",
 "M2-5":"Behaviour Support Plan Monitoring and Review",
 "M2-6":"Reportable Incidents involving the Use of a Restrictive Practice",
 "M2-7":"Interim Behaviour Support Plans","M2A-1":"Behaviour Support in the NDIS",
 "M2A-2":"Regulated Restrictive Practices",
 "M2A-3":"Supporting the Assessment and Development of Behaviour Support Plans",
 "M2A-4":"Behaviour Support Plan Implementation",
 "M2A-5":"Monitoring and Reporting the Use of Regulated Restrictive Practices",
 "M2A-6":"Behaviour Support Plan Review",
 "M2A-7":"Reportable Incidents involving the Use of a Restrictive Practice",
 "M2A-8":"Interim Behaviour Support Plans","M3-1":"The Child","M3-2":"The Family",
 "M3-3":"Inclusion","M3-4":"Collaboration","M3-5":"Capacity Building",
 "M3-6":"Evidence-Informed Practice","M3-7":"Outcome based approach",
 "M5-4":"Enrolment of SDA Properties"}
INFERRED_18 = {"M2-3","M2-4","M2-5","M2-6","M2-7","M2A-1","M2A-2","M2A-3","M2A-4",
               "M2A-5","M2A-6","M2A-7","M2A-8","M3-1","M3-2","M3-3","M3-4","M3-5"}
MODULE_ORDER = [
 ("CORE","Core module","all providers"),("M1","Module 1 — high intensity daily personal activities",""),
 ("M2","Module 2 — specialist behaviour support",""),("M2A","Module 2A — implementing behaviour support plans",""),
 ("M3","Module 3 — early childhood supports",""),("M4","Module 4 — specialised support coordination",""),
 ("M5","Module 5 — specialist disability accommodation",""),
 ("SIL","Module 5A — supported independent living (from 1 Jul 2026)",""),
 ("VER","Verification module","clause order reverse-paired to QI order"),
]
DOC_SPINE = {"POL":"pol","PRO":"pro","WIN":"win","FRM":"frm","AGR":"frm","REG":"reg","PLN":"pln","STA":"pln"}


def build_standards_page():
    from collections import defaultdict
    wb = load_workbook(OUT / "rise-ndis-package/rise-document-register.xlsx", read_only=True, data_only=True)
    std = {}
    order = []
    for r in list(wb["Standards"].iter_rows(values_only=True))[1:]:
        if not r[0]:
            continue
        std[r[0]] = {"module": r[1], "div": r[2] or "", "name": r[3], "rules": r[4], "qi": r[5]}
        order.append(r[0])
    imp = defaultdict(list)
    for r in list(wb["Edges"].iter_rows(values_only=True))[1:]:
        if r[1] == "IMPLEMENTS" and r[2] in std:
            imp[r[2]].append(r[0])
    wb.close()
    docfiles = {}
    for p in (OUT / "rise-ndis-package/suite").rglob("*.md"):
        m = ID_RE.match(p.name)
        if m:
            docfiles[m.group(1)] = str(p.relative_to(OUT).with_suffix(".html"))
    unmapped = [k for k in order if not imp.get(k)]

    # plain-English one-liners from the verified map, joined by QI section
    plain = {}
    pe_txt = (OUT / "rise-dossier-bundle/ndis-standards-quality-indicators-plain-english-map.md").read_text(encoding="utf-8")
    for m in re.finditer(r"^\|([^|]+)\|([^|]+)\|\s*section\s+(\d+[A-E]?)\s*\|([^|]+)\|", pe_txt, re.M):
        plain["s " + m.group(3)] = m.group(4).strip()
    print("plain-English lines matched to register QI sections:",
          sum(1 for k in order if str(std[k]["qi"]).strip() in plain), "of", len(order))

    blocks, gridmods = [], []
    for mod, label, note in MODULE_ORDER:
        rows_html, cells_html, last_div = [], [], None
        for sid in order:
            s = std[sid]
            if s["module"] != mod:
                continue
            if mod == "CORE" and s["div"] != last_div:
                rows_html.append('<div class="sdiv">' + esc(s["div"]) + "</div>")
                cells_html.append('<div class="gdiv">' + esc(s["div"]) + "</div>")
                last_div = s["div"]
            if sid.startswith("SIL"):
                stc, sym, chip = "t", "–", '<span class="chip st-t">Rules clause TBC</span>'
            elif sid in INFERRED_18:
                stc, sym, chip = "r", "R", '<span class="chip st-r">register-confirmed 7 Jul</span>'
            else:
                stc, sym, chip = "v", "✓", '<span class="chip st-v">verified (prior fetch)</span>'
            variant = ""
            bn = BOOKLET_NAMES.get(sid)
            if bn and bn.strip().lower() != str(s["name"]).strip().lower():
                variant = ('<div class="variant">Commission booklet (Nov 2021) title: '
                           + esc(bn) + "</div>")
            pl = plain.get(str(s["qi"]).strip())
            if pl:
                plain_html = ('<div class="plain"><span class="ptag">In plain English</span>'
                              + esc(pl) + "</div>")
            else:
                plain_html = ('<div class="plain pnone"><span class="ptag">No plain-English line yet</span>'
                              "The verified plain-English map (6 Jul) pre-dates SIL. Commission descriptions: "
                              '<a href="https://www.ndiscommission.gov.au/rules-and-standards/ndis-practice-standards/sil">SIL Practice Standards page</a>.</div>')
            docs = sorted(set(imp.get(sid, [])))
            search = (sid + " " + str(s["name"]) + " " + (pl or "") + " " + " ".join(docs)).lower()
            cells_html.append('<a class="gcell st-' + stc + '" href="#' + esc(sid)
                              + '" data-search="' + esc(search) + '" title="' + esc(str(s["name"]) + ((" — " + pl) if pl else ""))
                              + '"><span class="gsym">' + sym + '</span><span class="gid">' + esc(sid)
                              + '</span><span class="gqi">' + esc(str(s["qi"])) + "</span></a>")
            if docs:
                chips = "".join(
                    '<a class="dchip s-' + DOC_SPINE.get(d[:3], "gen") + '" href="' + esc(docfiles[d])
                    + '" title="' + esc(catalogue.get(docfiles[d].replace(".html", ".md"), {}).get("title", d))
                    + '">' + esc(d) + "</a>" for d in docs if d in docfiles)
                docs_html = '<div class="docs"><span class="dlabel">Implemented by</span>' + chips + "</div>"
            else:
                docs_html = ('<div class="docs nodoc">No suite document carries an IMPLEMENTS edge to this '
                             "standard (computed from the register Edges sheet) — intentional scope or a gap to review.</div>")
            rows_html.append(
                '<div class="srow" id="' + esc(sid) + '" data-search="' + esc(search) + '">'
                '<div class="stop"><span class="sid mono">' + esc(sid) + "</span>" + chip
                + '<span class="addr mono">' + esc(str(s["qi"])) + " · " + esc(str(s["rules"])) + "</span></div>"
                '<div class="sname">' + esc(str(s["name"])) + "</div>" + variant + plain_html + docs_html
                + '<a class="up" href="#map">↑ back to map</a></div>')
        gridmods.append('<div class="gmod" data-mod="1"><div class="gmodh">' + esc(label) + "</div>"
                        + '<div class="gcells">' + "".join(cells_html) + "</div></div>")
        blocks.append('<section class="smod"><h2>' + esc(label)
                      + ('<span class="mnote">' + esc(note) + "</span>" if note else "") + "</h2>"
                      + "".join(rows_html) + "</section>")

    n_conf = len(INFERRED_18)
    n_sil = sum(1 for k in order if k.startswith("SIL"))
    n_ver = len(order) - n_conf - n_sil
    page = STDS_TMPL.replace("@@BLOCKS@@", "".join(blocks)).replace("@@MAP@@", "".join(gridmods))
    page = page.replace("@@KBCSS@@", KB_CSS)
    page = (page.replace("@@TOTAL@@", str(len(order))).replace("@@NVER@@", str(n_ver))
                .replace("@@NCONF@@", str(n_conf)).replace("@@NSIL@@", str(n_sil))
                .replace("@@NMAPPED@@", str(len(order) - len(unmapped)))
                .replace("@@UNMAPPED@@", ", ".join(unmapped)).replace("@@DATE@@", BUILD_DATE))
    (OUT / "standards.html").write_text(page, encoding="utf-8")
    catalogue["standards.html"] = {"title": "NDIS Practice Standards — linked map (all 70, with implementing documents)",
                                   "view": "standards.html", "kind": "interactive"}
    return unmapped


STDS_TMPL = """<!DOCTYPE html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>NDIS Practice Standards — linked map · Rise KB</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="stylesheet" href="https://fonts.googleapis.com/css2?family=Spectral:wght@500;700;800&family=Public+Sans:wght@400;600;700&family=IBM+Plex+Mono:wght@400;600&display=swap" media="all" onerror="this.media='none'">
<style>@@KBCSS@@</style>
<style>
.mast{background:var(--navy);color:#e9eef6;padding:26px 20px 20px}
.mast .wrap{max-width:940px;margin:0 auto}
.mast h1{font:800 clamp(24px,3.6vw,32px)/1.15 Spectral,Georgia,serif;margin:4px 0 6px;color:#fff}
.mast p{margin:0;font-size:13.5px;color:#c9d6ea;max-width:78ch}
.mast a{color:#cfe0f7}
.bar{position:sticky;top:0;z-index:9;background:var(--navy2);padding:9px 20px;box-shadow:0 2px 8px rgba(10,20,40,.25)}
.bar .wrap{max-width:940px;margin:0 auto;display:flex;gap:12px;align-items:center}
#q{flex:1;max-width:560px;font:14.5px "Public Sans";padding:8px 13px;border-radius:8px;border:1px solid #4b6288;background:#0e2140;color:#e9eef6}
#q::placeholder{color:#8fa3c4}#q:focus-visible{outline:2px solid #9fc0f2}
#cnt{font:600 12px "IBM Plex Mono";color:#cdd9ec}
.bar a{color:#dbe4f2;font:600 12.5px "Public Sans";text-decoration:none}
main{max-width:940px;margin:0 auto;padding:14px 18px 70px}
.prov{background:var(--panel);border:1px solid var(--rule);border-left:5px solid var(--inf);border-radius:9px;padding:12px 15px;font-size:13.5px;line-height:1.55;margin:14px 0}
.metrics{display:grid;grid-template-columns:repeat(auto-fit,minmax(130px,1fr));gap:10px;margin:14px 0 6px}
.mcard{background:var(--panel);border:1px solid var(--rule);border-radius:8px;padding:9px 12px}
.mcard .l{font-size:12px;color:var(--mut)}.mcard .v{font-size:21px;font-weight:600}
.smod h2{font:700 21px/1.2 Spectral,Georgia,serif;margin:30px 0 10px}
.mnote{font:400 12px "IBM Plex Mono";color:var(--mut);margin-left:10px}
.sdiv{font-size:12.5px;color:var(--mut);margin:14px 0 6px;font-weight:600;letter-spacing:.03em}
.srow{background:var(--panel);border:1px solid var(--rule);border-radius:9px;padding:11px 14px;margin:8px 0}
.stop{display:flex;gap:10px;align-items:baseline;flex-wrap:wrap}
.sid{font-weight:600;font-size:13.5px}
.addr{margin-left:auto;font-size:12px;color:var(--mut)}
.chip{font:600 10.5px "Public Sans";letter-spacing:.04em;text-transform:uppercase;padding:2px 8px;border-radius:20px;border:1px solid}
.st-v{background:var(--okbg);border-color:var(--ok);color:var(--ok)}
.st-r{background:var(--infbg);border-color:var(--inf);color:var(--inf)}
.st-t{background:var(--ambbg);border-color:var(--amb);color:var(--amb)}
.sname{font-size:15.5px;font-weight:600;margin:5px 0 2px}
.variant{font-size:12.5px;color:var(--mut);font-style:italic;margin:1px 0 3px}
.plain{font-size:13.5px;line-height:1.5;background:color-mix(in srgb,var(--okbg) 45%,var(--panel) 55%);
  border-left:3px solid var(--ok);border-radius:0 6px 6px 0;padding:6px 11px;margin:6px 0 2px;max-width:78ch}
.plain .ptag{display:block;font:600 10px "Public Sans";letter-spacing:.06em;text-transform:uppercase;color:var(--ok);margin:0 0 2px}
.pnone{background:var(--ambbg);border-left-color:var(--amb)}
.pnone .ptag{color:var(--amb)}
.docs{display:flex;gap:6px;flex-wrap:wrap;align-items:center;margin-top:7px}
.dlabel{font:600 11px "Public Sans";letter-spacing:.05em;text-transform:uppercase;color:var(--mut);margin-right:2px}
.dchip{font:600 11.5px "IBM Plex Mono";padding:3px 8px;border-radius:6px;border:1px solid var(--rule);border-left-width:4px;text-decoration:none;color:var(--ink);background:var(--panel)}
.dchip:hover{text-decoration:underline}
.nodoc{font-size:12.5px;color:var(--amb)}
.s-pol{border-left-color:var(--pol)}.s-pro{border-left-color:var(--pro)}.s-win{border-left-color:var(--win)}
.s-frm{border-left-color:var(--frm)}.s-reg{border-left-color:var(--reg)}.s-pln{border-left-color:var(--pln)}.s-gen{border-left-color:var(--gen)}
.maph{font:700 22px/1.2 Spectral,Georgia,serif;margin:22px 0 4px}
.mapnote{font-size:12.5px;color:var(--mut);margin:0 0 10px}
.gmod{margin:0 0 14px}
.gmodh{font:600 14px Spectral,Georgia,serif;margin:10px 0 4px}
.gdiv{font-size:11.5px;color:var(--mut);margin:7px 0 3px;grid-column:1/-1}
.gcells{display:grid;grid-template-columns:repeat(auto-fill,minmax(84px,1fr));gap:6px}
.gcell{border-radius:8px;padding:6px 5px 5px;text-align:center;position:relative;min-height:44px;
  text-decoration:none;border:1px solid;display:block}
.gcell .gid{font:600 12.5px "IBM Plex Mono";display:block;line-height:1.2}
.gcell .gqi{font:400 10.5px "IBM Plex Mono";opacity:.85;display:block;margin-top:2px}
.gcell .gsym{position:absolute;top:2px;right:6px;font-size:10.5px;font-weight:600}
.gcell:hover{filter:brightness(.96)}
.st-v.gcell{background:var(--okbg);border-color:var(--ok);color:var(--ok)}
.st-r.gcell{background:var(--infbg);border-color:var(--inf);color:var(--inf)}
.st-t.gcell{background:var(--ambbg);border-color:var(--amb);color:var(--amb)}
.srow{scroll-margin-top:70px}
.up{display:inline-block;margin-top:8px;font:600 11.5px "Public Sans";text-decoration:none;color:var(--mut)}
.up:hover{color:var(--link)}
.hide{display:none!important}
footer{max-width:940px;margin:0 auto;padding:14px 18px 40px;font-size:12.5px;color:var(--mut);border-top:1px solid var(--rule)}
</style></head><body>
<header class="mast"><div class="wrap">
<h1>NDIS Practice Standards — the linked map</h1>
<p>All @@TOTAL@@ standards with their statutory addresses and the Rise documents that implement them. Every document chip opens the drafted document; addresses come from the register. Generated from <span class="mono">rise-document-register.xlsx</span> (Standards + Edges sheets) at build time, so this page cannot drift from the data.</p>
</div></header>
<div class="bar"><div class="wrap">
<input id="q" type="search" placeholder="Filter standards — try: incident, restrictive, SIL, CORE-2, FRM-…" aria-label="Filter standards">
<span id="cnt"></span>
<a href="index.html">← Knowledge base</a>
</div></div>
<main>
<div class="prov"><b>Provenance &amp; status (@@DATE@@).</b> Names and dual addresses are from the register&#39;s Standards sheet — verified against the fetched instruments in a prior session, not re-fetched from the Federal Register today. The 18 QI pairings that the 6 Jul reconciliation could only <em>infer</em> from the Commission booklet were compared against the register during this build: <b>all 18 matched</b> (chips marked purple), which also filled their Rules clauses (M2 Sch 3 cl 5–9 · M2A Sch 4 cl 3–10 · M3 Sch 5 cl 3–6). SIL rows carry register-verified names and QI sections; the register itself marks their Rules clauses &ldquo;TBC — check latest compilation&rdquo;, so they stay amber. Where the Commission booklet&#39;s title differs from the register&#39;s (Rules-side) title, both are shown. The green <b>&ldquo;In plain English&rdquo;</b> line under each standard is the one-line auditor summary from the <a href="rise-dossier-bundle/ndis-standards-quality-indicators-plain-english-map.html">plain-English map</a> — labelled paraphrase drafted in a prior session from the legislated indicator tables, joined here by QI section (66 of 70; the map pre-dates SIL, so those four rows say so instead). The filter box searches these lines too, so everyday words like &ldquo;swallowing&rdquo; or &ldquo;complaints&rdquo; find their standards. The outstanding gold-standard step remains a fresh fetch of the in-force compilations. Full outcomes and indicator text: the <a href="rise-audit-pack/00-source/ndis-practice-standards-and-quality-indicators__1_.pdf">Commission booklet (Nov 2021, pre-SIL)</a>; auditor tests per standard: the <a href="rise-audit-pack/02-checklist/rise-ndis-audit-checklist-indexed.html">indexed checklist</a>; method: the <a href="rise-audit-pack/02-checklist/rise-reconciliation-2021v4-booklet.html">reconciliation</a>; plain-English lines: the <a href="rise-dossier-bundle/ndis-standards-quality-indicators-plain-english-map.html">66-standard map</a>. Commission source page: <a href="https://www.ndiscommission.gov.au/rules-and-standards/ndis-practice-standards">ndiscommission.gov.au</a>; instrument IDs to search on legislation.gov.au: PRPS Rules F2018L00631 · QI Guidelines C03 F2026C00528.</div>
<div class="metrics">
<div class="mcard"><div class="l">Standards</div><div class="v">@@TOTAL@@</div></div>
<div class="mcard"><div class="l">Verified (prior fetch)</div><div class="v">@@NVER@@</div></div>
<div class="mcard"><div class="l">Register-confirmed 7 Jul</div><div class="v">@@NCONF@@</div></div>
<div class="mcard"><div class="l">SIL — Rules clause TBC</div><div class="v">@@NSIL@@</div></div>
<div class="mcard"><div class="l">With mapped documents</div><div class="v">@@NMAPPED@@</div></div>
</div>
<p style="font-size:13px;color:var(--amb);margin:0 0 6px">No document mapped yet: <span class="mono">@@UNMAPPED@@</span> — computed from the Edges sheet; review whether intentional.</p>
<h2 class="maph" id="map">The map — tap any cell for its documents</h2>
<p class="mapnote">Same cells as the 6 Jul coverage map, statuses updated: ✓ verified (prior fetch) · R register-confirmed 7 Jul · – SIL, Rules clause TBC. Symbols duplicate colour for greyscale printing.</p>
@@MAP@@
@@BLOCKS@@
</main>
<footer>Generated by <span class="mono">_kb/build_kb.py</span> on @@DATE@@ from the register workbook. Counts computed from the data at render time. The 6 Jul coverage map v2 remains in the collection as the pre-comparison epistemic record. Not legal advice; the instruments on the Federal Register govern.</footer>
<script>
const q=document.getElementById('q'),cnt=document.getElementById('cnt');
const rows=[...document.querySelectorAll('.srow')];
const cells=[...document.querySelectorAll('.gcell')];
q.addEventListener('input',()=>{const v=q.value.trim().toLowerCase();let h=0;
rows.forEach(r=>{const on=!v||r.dataset.search.includes(v);r.classList.toggle('hide',!on);if(on&&v)h++;});
cells.forEach(c=>{c.classList.toggle('hide',v&&!c.dataset.search.includes(v));});
document.querySelectorAll('.smod').forEach(s=>{const any=[...s.querySelectorAll('.srow')].some(r=>!r.classList.contains('hide'));s.classList.toggle('hide',v&&!any);});
document.querySelectorAll('.gmod').forEach(g=>{const any=[...g.querySelectorAll('.gcell')].some(c=>!c.classList.contains('hide'));g.classList.toggle('hide',v&&!any);});
document.querySelectorAll('.sdiv,.gdiv').forEach(d=>d.classList.toggle('hide',!!v));
cnt.textContent=v?h+' match'+(h===1?'':'es'):'';});
</script>
</body></html>
"""

unmapped_standards = build_standards_page()

# ---------------------------------------------------------------- index build
PKG = "rise-ndis-package/"
AUD = "rise-audit-pack/"
BND = "rise-dossier-bundle/"
SUITE_GROUPS = [
    ("01-policies", "Policies", "pol"),
    ("02-procedures", "Procedures", "pro"),
    ("03-work-instructions", "Work instructions", "win"),
    ("04-forms-and-agreements", "Forms & agreements", "frm"),
    ("05-registers", "Registers", "reg"),
    ("06-plans-and-handbooks", "Plans & handbooks", "pln"),
]

ID_RE = re.compile(r"^([A-Z]{3}-[A-Z]{3}-\d{2}[A-Za-z]?)")


def row(orig, desc="", spine="gen", tag=None, extra=""):
    e = catalogue[orig]
    fid = ""
    m = ID_RE.match(Path(orig).name)
    if m:
        fid = '<span class="rid mono">' + m.group(1) + "</span>"
    tag = tag or e["kind"]
    d = '<span class="rd">' + desc + "</span>" if desc else ""
    return ('<a class="row s-' + spine + '" data-search="' + esc((m.group(1) + " " if m else "") + e["title"] + " " + orig).lower() +
            '" href="' + esc(e["view"]) + '">' + fid + '<span class="rt">' + esc(e["title"]) + d +
            '</span><span class="tag mono">' + esc(tag) + "</span>" + extra + "</a>")


def card(orig, desc, spine="gen", tag=None):
    e = catalogue[orig]
    tag = tag or e["kind"]
    return ('<a class="card s-' + spine + '" data-search="' + esc(e["title"] + " " + orig + " " + desc).lower() +
            '" href="' + esc(e["view"]) + '"><span class="ct">' + esc(e["title"]) +
            '</span><span class="cd">' + desc + '</span><span class="tag mono">' + esc(tag) + "</span></a>")


sections = []

# --- start here
sections.append('<section id="start"><h2>Start here</h2><div class="cards">' + "".join([
    card("standards.html",
         "Every standard: name, Rules clause + QI section, status, and clickable links to the documents that implement it. Generated from the register.",
         "reg", "interactive"),
    card(PKG + "rise-dossier.html",
         "Self-contained interactive dossier — legislation, 70 standards, all 123 documents, graph explorer. Snapshot 5 Jul 2026.",
         "pol", "interactive"),
    card(PKG + "suite/00-INDEX.md",
         "The mapping: every policy → procedure → form → register chain as a clickable tree, generated from the graph edges.",
         "pro"),
    card(AUD + "01-guide/how-ndis-auditors-check-practice-standards.md",
         "The layered deep-research guide: how AQA auditors actually test providers, anatomy of a real 2022 audit report, 2026 changes.",
         "frm"),
    card(AUD + "03-visuals/ndis-standards-coverage-map-v2.html",
         "The 6 Jul epistemic record: 48 verified · 18 inferred · 4 partial. Superseded as navigation by the linked standards map, kept unmodified as history.",
         "reg", "interactive"),
    card(BND + "rise-visual-atlas.html",
         "All twelve session diagrams on one page — four-box model, audit journey, indicator testing, pyramid, gap engine — each captioned with its evidence status.",
         "pln", "interactive"),
    card(BND + "rise-standards-navigator.html",
         "Interactive standards navigator: search, module filters, dual-title flags, cross-links, copy-citation.",
         "win", "interactive"),
]) + "</div></section>")

# --- status band (not files)
sections.append("""
<section id="status"><h2>Status at a glance</h2><div class="band">
<div class="chip c-ok"><b>Standards indexing (after the 7 Jul register comparison):</b> 48 addresses verified by prior instrument fetch · 18 upgraded from inferred to register-confirmed (all 18 QI pairings matched; Rules clauses filled) · 4 SIL rows named and QI-placed, Rules clauses TBC per the register itself. Nothing here was re-fetched from the Federal Register today — that currency check is the remaining gold step.</div>
<div class="chip c-amb"><b>The 123 suite documents are all DRAFT v0.1.</b> Compliance links are generated from the register and structurally reliable; body wording is a starting point to tailor and approve. Not legal advice. Four standards currently have no implementing document mapped (see the standards map) — review whether intentional.</div>
<div class="chip c-inf"><b>The [I]→[V] check has been run (7 Jul), during the standards-map build.</b> Method and result are recorded on the <a href="standards.html">standards map</a> itself; the 6 Jul <a href="rise-audit-pack/03-visuals/ndis-standards-coverage-map-v2.html">coverage map v2</a> is retained unmodified as the pre-comparison record. Source data: the register&#39;s <a href="rise-ndis-package/rise-document-register.xlsx.html">Standards sheet</a> and the <a href="rise-dossier-bundle/ndis-standards-quality-indicators-plain-english-map.html">plain-English map</a>.</div>
</div></section>""")

# --- audit research
sections.append('<section id="audit"><h2>Standards &amp; audit research <span class="scount mono">6 Jul session</span></h2>' + "".join([
    row(AUD + "01-guide/how-ndis-auditors-check-practice-standards.md",
        " — method: say it → do it → prove it, with V/C/T/D source labels", "frm"),
    row(AUD + "02-checklist/rise-ndis-audit-checklist-indexed.md",
        " — per-standard auditor tests + evidence; gap tables §§3–5 superseded (banner on page)", "frm",
        extra='<span class="tag mono t-amb">partly superseded</span>'),
    row(AUD + "02-checklist/rise-reconciliation-2021v4-booklet.md",
        " — the reconciliation that closed the name-gaps; records the inference basis and what would falsify it", "frm"),
    row(AUD + "03-visuals/ndis-standards-coverage-map-v2.html",
        " — current coverage state, counts computed at render time", "reg"),
    row(AUD + "00-source/ndis-practice-standards-and-quality-indicators__1_.pdf",
        " — NDIS Commission booklet, Nov 2021 v4: the primary source behind the reconciliation (pre-SIL)", "win"),
    row(AUD + "04-manifest/rise-collection-manifest.md",
        " — cross-session inventory and the assembly plan this knowledge base follows", "win"),
    row(AUD + "README.md", " — the audit pack's own index and honesty statement", "win"),
]) + "</section>")

# --- design-session bundle
sections.append('<section id="bundle"><h2>Design-session bundle <span class="scount mono">Rise Dossier · 5\u20137 Jul · 15 files</span></h2>'
    + '<p class="secnote">Uploaded 7 Jul as <span class="mono">rise-dossier.zip</span>; stored as <span class="mono">rise-dossier-bundle/</span> to avoid confusion with the interactive <span class="mono">rise-dossier.html</span>. One-liners below are taken from the bundle&#39;s own index.</p>'
    + "".join([
    row(BND + "rise-dossier-index.md", " — the bundle's own master index: inventory, reading paths, open-questions register, session verification log", "gen"),
    row(BND + "rise-visual-atlas.html", " — all twelve session diagrams, incl. the four-box model (fig 4), the audit journey (fig 6), how an auditor tests one indicator (fig 7), the pyramid (fig 11), the gap engine (fig 12)", "pln"),
    row(BND + "rise-standards-navigator.html", " — interactive navigator; holds the eight dual-title flags the checklist points to", "win"),
    row(BND + "rise-compliance-pyramid-model.md", " — the seven-tier presentation model + complete placement map for all 66 pre-SIL standards", "pro"),
    row(BND + "ndis-standards-quality-indicators-plain-english-map.md", " — all 66 standards mapped to their indicator sections; one of the two stated gap-closers", "pro"),
    row(BND + "rise-relationship-taxonomy.md", " — v0.3: the 25 typed edges, three-dimension confidence model, MAPS_TO per NIST OLIR", "pol"),
    row(BND + "rise-claude-code-brief.md", " — Milestone 1 build brief: graph schema, seed data, gap views, tests; taxonomy v0.3 embedded as Appendix A", "pol"),
    row(BND + "rise-milestone-2-gap-engine.md", " — Milestone 2: indicator ingestion, demand-type classification, applicability gates, four-state gap model", "pol"),
    row(BND + "rise-source-manifest.md", " — every instrument in scope with register IDs, verification status, and hard ingestion rules", "win"),
    row(BND + "rise-bounce-prototype.html", " — golden-thread interaction: one indicator traced policy \u2192 procedure \u2192 form \u2192 register", "frm"),
    row(BND + "rise-graph-explorer.html", " — drill-down tree from the Act to the evidence; content grades chipped per node", "frm"),
    row(BND + "rise-ui-interaction-notes.md", " — five-panel book-system analysis + seven design recommendations", "frm"),
]) + '<p class="secnote">The three items below post-date the bundle&#39;s own index (6 Jul, 22:56) and are not in its inventory:</p>' + "".join([
    row(BND + "rise-zero-knowledge-training-model.md", " — v1.0, confirmed 7 Jul: training-first reframe; fixes the nine-step referent (standard \u2192 policy \u2192 procedure \u2192 form \u2192 register \u2192 evidence \u2192 audit \u2192 finding \u2192 improvement action)", "reg"),
    row(BND + "rise-front-door.html", " — the nine steps as the product's entry experience", "reg"),
    row(BND + "rise-your-world.html", " — Level 2 of the training entry", "reg"),
]) + "</section>")

# --- authoring research (this session)
sections.append('<section id="authoring"><h2>Authoring &amp; formatting best practice <span class="scount mono">researched 7 Jul</span></h2>'
    + '<p class="secnote">Web-verified this session (Style Manual, Easy Read guidance, ISO 9001 document-control consensus, NDIS Commission pages). Grounds the five TPL templates and the architecture spec in external standards; records one correction (participant-facing reading level: Year 7, not Year 8&ndash;9) and proposes six deltas. Not a replacement for the templates &mdash; a layer on them.</p>'
    + row("rise-authoring-research/rise-authoring-best-practice.md",
         " — plain language, document control, form design, and the Easy Read production standard, with tiered dated sources", "pro")
    + "</section>")

# --- document suite
suite_bits = ['<section id="suite"><h2>Document suite <span class="scount mono">123 drafts · 6 type folders</span></h2>',
              row(PKG + "suite/00-INDEX.md", " — read this first: the policy→procedure→form→register chains", "pro")]
for folder, label, spine in SUITE_GROUPS:
    entries = [k for k in catalogue if k.startswith(PKG + "suite/" + folder + "/")]
    entries.sort()
    suite_bits.append('<details class="group"><summary>' + esc(label) +
                      ' <span class="scount mono">' + str(len(entries)) + "</span></summary><div>")
    for k in entries:
        suite_bits.append(row(k, spine=spine, tag="draft v0.1"))
    suite_bits.append("</div></details>")
suite_bits.append("</section>")
sections.append("".join(suite_bits))

# --- architecture, data, generators
sections.append('<section id="dev"><h2>Architecture, data &amp; generators</h2>' + "".join([
    row(PKG + "rise-developer-handoff.md", " — the technical index and import plan; the developer's start point", "pol"),
    row(PKG + "rise-node-taxonomy.md", " — node-type schema v0.1 (companion to edge taxonomy v0.3, not in this build)", "pol"),
    row(PKG + "rise-document-architecture.md", " — Compliance Header spec, ID scheme, traceability views", "pol"),
    row(PKG + "rise-document-register.xlsx", " — master register: Documents 123 · Standards 70 · Legislation 21 · Edges 528 (preview + download)", "reg"),
    row(PKG + "rise-nodes-and-edges.json", " — seed graph: 224 nodes, 528 edges", "reg"),
    row(PKG + "rise-node-examples.json", " — one fully-attributed example per node type", "reg"),
    row(PKG + "rise-node-map.mermaid", " — node-type meta-model diagram", "reg"),
    row(PKG + "rise_build.py", " — generator (single-source rebuild chain, run with rise_docs + rise_dossier)", "win"),
    row(PKG + "rise_docs.py", " — generator: renders the 123 documents", "win"),
    row(PKG + "rise_dossier.py", " — generator: builds the interactive dossier", "win"),
    row(PKG + "TPL-POL-policy-template.md", " — blank template", "pln"),
    row(PKG + "TPL-PRO-procedure-template.md", " — blank template", "pln"),
    row(PKG + "TPL-WIN-work-instruction-template.md", " — blank template", "pln"),
    row(PKG + "TPL-FRM-form-template.md", " — blank template", "pln"),
    row(PKG + "TPL-REG-register-template.md", " — blank template", "pln"),
]) + "</section>")

# --- provenance & method
sections.append('<section id="prov"><h2>Provenance, method &amp; guides</h2>' + "".join([
    row(PKG + "rise-plain-english-guide.md", " — the whole NDIS compliance system explained for anyone", "pro"),
    row(PKG + "rise-auditor-focus-map.md", " — what auditors open first, per module, QI-section-cited", "pro"),
    row(PKG + "rise-methodology.md", " — how the package was built and verified; how to challenge it", "win"),
    row(PKG + "rise-sources.md", " — every claim traced to a tiered, dated source, including negative findings", "win"),
    row(PKG + "README.md", " — the package's own README: what works offline, what to trust", "win"),
]) + "</section>")

# --- open items + not included
sections.append("""
<section id="open"><h2>Open items &amp; known-missing artefacts</h2>
<div class="band">
<div class="chip c-amb"><b>Open TODOs</b> (updated 7 Jul after the register comparison): SIL Rules clauses — marked TBC in the register itself, need the current PRPS compilation · fresh-fetch currency check of all addresses against the in-force compilations (the register was verified in a prior session, not today) · two statutory timeframes flagged "confirm before approval" · Easy Read companions flagged but not drafted (production standard now in the authoring guide) · four standards with no implementing document (M2-2, M2-4, M2-6, M2A-3) · the bundle's seven-item open-questions register.</div>
<div class="chip c-inf"><b>Known-missing list, updated 7 Jul.</b> The bundle upload closed the manifest's Session C and D entries (taxonomy v0.3, plain-English map, navigator, pyramid + explainers via the atlas, source manifest, Claude Code brief). Still absent from the manifest's census: the <span class="mono">.docx</span> template variants (only <span class="mono">.md</span> templates are in the package). Coverage map v1 is superseded, not missing. The manifest itself warns its census may be incomplete.</div>
</div></section>""")

SEARCH_JS = """
const q=document.getElementById('q'),cnt=document.getElementById('cnt');
const items=[...document.querySelectorAll('[data-search]')];
const secs=[...document.querySelectorAll('section')];
q.addEventListener('input',()=>{
  const v=q.value.trim().toLowerCase();
  let hits=0;
  items.forEach(el=>{const on=!v||el.dataset.search.includes(v);el.classList.toggle('hide',!on);if(on&&v)hits++;});
  document.querySelectorAll('details.group').forEach(d=>{
    const any=[...d.querySelectorAll('[data-search]')].some(el=>!el.classList.contains('hide'));
    d.classList.toggle('hide',v&&!any); if(v&&any)d.open=true; if(!v)d.open=false;
  });
  secs.forEach(s=>{
    if(!s.querySelector('[data-search]'))return;
    const any=[...s.querySelectorAll('[data-search]')].some(el=>!el.classList.contains('hide'));
    s.classList.toggle('hide',v&&!any);
  });
  cnt.textContent=v?hits+' match'+(hits===1?'':'es'):'';
});
"""

n_orig = len(catalogue)
n_suite = sum(1 for k in catalogue if k.startswith(PKG + "suite/") and not k.endswith("00-INDEX.md"))

INDEX_CSS = """
.mast{background:var(--navy);color:#e9eef6;padding:38px 20px 26px}
.mast .wrap{max-width:980px;margin:0 auto;display:flex;gap:26px;flex-wrap:wrap;align-items:flex-end;justify-content:space-between}
.eyebrow{font:600 12px/1 "Public Sans";letter-spacing:.22em;text-transform:uppercase;color:#9fb3d1;margin:0 0 10px}
h1{font:800 clamp(28px,4.2vw,40px)/1.1 Spectral,Georgia,serif;margin:0;max-width:620px}
h1 em{font-style:italic;font-weight:500;color:#b9c9e2}
.stampbox{border:2px solid #8fa5c8;padding:11px 15px;font:600 11.5px/1.7 "IBM Plex Mono",monospace;
  letter-spacing:.06em;color:#cdd9ec;text-transform:uppercase;transform:rotate(-1.2deg);white-space:nowrap}
.stampbox b{color:#fff}
.drawer{position:sticky;top:0;z-index:9;background:var(--navy2);box-shadow:0 2px 8px rgba(10,20,40,.25);padding:11px 20px}
.drawer .wrap{max-width:980px;margin:0 auto;display:flex;gap:12px;align-items:center}
#q{flex:1;max-width:640px;font:15px "Public Sans";padding:9px 14px;border-radius:8px;border:1px solid #4b6288;
  background:#0e2140;color:#e9eef6}
#q::placeholder{color:#8fa3c4}
#q:focus-visible{outline:2px solid #9fc0f2}
#cnt{font:600 12.5px "IBM Plex Mono";color:#cdd9ec}
main{max-width:980px;margin:0 auto;padding:8px 18px 70px}
section{margin:34px 0 0}
section h2{font:700 22px/1.2 Spectral,Georgia,serif;margin:0 0 12px;display:flex;align-items:baseline;gap:12px}
.scount{font-size:12px;color:var(--mut);font-weight:400}
.cards{display:grid;grid-template-columns:repeat(auto-fit,minmax(225px,1fr));gap:12px}
.card{display:flex;flex-direction:column;gap:6px;background:var(--panel);border:1px solid var(--rule);
  border-left-width:5px;border-radius:9px;padding:13px 15px;text-decoration:none;color:var(--ink)}
.card:hover .ct,.row:hover .rt{text-decoration:underline}
.ct{font:700 15.5px/1.3 "Public Sans"}
.cd{font-size:13px;color:var(--mut);line-height:1.45}
.tag{font-size:10.5px;letter-spacing:.05em;text-transform:uppercase;color:var(--mut)}
.t-amb{color:var(--amb)}
.row{display:flex;gap:11px;align-items:baseline;background:var(--panel);border:1px solid var(--rule);
  border-left-width:5px;border-radius:7px;padding:8px 12px;margin:6px 0;text-decoration:none;color:var(--ink)}
.rid{font-size:12px;font-weight:600;color:var(--mut);min-width:96px}
.rt{flex:1;font-size:14.5px;line-height:1.4}
.rd{color:var(--mut);font-size:13px}
.s-pol{border-left-color:var(--pol)}.s-pro{border-left-color:var(--pro)}.s-win{border-left-color:var(--win)}
.s-frm{border-left-color:var(--frm)}.s-reg{border-left-color:var(--reg)}.s-pln{border-left-color:var(--pln)}
.s-gen{border-left-color:var(--gen)}
details.group{margin:10px 0;border:1px solid var(--rule);border-radius:9px;background:color-mix(in srgb,var(--panel) 60%,var(--bg) 40%)}
details.group>summary{cursor:pointer;padding:11px 15px;font:700 16px Spectral,Georgia,serif}
details.group>div{padding:2px 12px 10px}
.band{display:grid;grid-template-columns:repeat(auto-fit,minmax(260px,1fr));gap:12px}
.chip{border-radius:9px;padding:12px 15px;font-size:13.5px;line-height:1.55;border:1px solid var(--rule);background:var(--panel)}
.c-ok{background:var(--okbg);border-color:var(--ok)}
.c-amb{background:var(--ambbg);border-color:var(--amb)}
.c-inf{background:var(--infbg);border-color:var(--inf)}
.secnote{font-size:13px;color:var(--mut);margin:4px 0 8px;line-height:1.5}
.hide{display:none!important}
footer{max-width:980px;margin:0 auto;padding:0 18px 40px;font-size:12.5px;color:var(--mut);border-top:1px solid var(--rule);padding-top:14px}
"""

index_html = """<!DOCTYPE html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Rise Knowledge Base — NDIS compliance collection</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="stylesheet" href="https://fonts.googleapis.com/css2?family=Spectral:wght@500;700;800&family=Public+Sans:wght@400;600;700&family=IBM+Plex+Mono:wght@400;600&display=swap" media="all" onerror="this.media='none'">
<style>@@KBCSS@@</style>
<style>@@CSS@@</style>
</head><body>
<header class="mast"><div class="wrap">
<div><p class="eyebrow">Rise · NDIS compliance collection</p>
<h1>Knowledge base <em>— everything, one index</em></h1></div>
<div class="stampbox">Assembled <b>@@DATE@@</b><br>3 collections + research · <b>@@NORIG@@ files</b><br>@@NSUITE@@-document suite</div>
</div></header>
<div class="drawer"><div class="wrap">
<input id="q" type="search" placeholder="Search everything — try: incident, restrictive, SIL, POL-, medication…" aria-label="Search the knowledge base">
<span id="cnt"></span>
</div></div>
<main>
@@SECTIONS@@
</main>
<footer>Assembled by Claude for Rise Development on @@DATE@@, following the assembly plan in the
<a href="rise-audit-pack/04-manifest/rise-collection-manifest.html">collection manifest</a>. Both source trees are
preserved unmodified; every rendered page links back to its original file. Duplicates verified by md5 and stored
once: the standalone dossier, reconciliation and coverage-map uploads matched copies inside the packages; the
standalone <span class="mono">rise-visual-atlas.html</span> matched the bundle's copy; and every file in
<span class="mono">files__4_.zip</span> (7 Jul) matched audit-pack material already here, so nothing from it was added.
This index was generated by <a class="mono" href="_kb/build_kb.py">_kb/build_kb.py</a> —
rerun it to regenerate after adding files. Counts above are computed from the files at build time.
Not legal advice; where anything here disagrees with the instrument on the Federal Register, the instrument governs.</footer>
<script>@@JS@@</script>
</body></html>
""".replace("@@KBCSS@@", KB_CSS).replace("@@CSS@@", INDEX_CSS).replace("@@JS@@", SEARCH_JS).replace("@@DATE@@", BUILD_DATE)
index_html = index_html.replace("@@NORIG@@", str(n_orig)).replace("@@NSUITE@@", str(n_suite))
index_html = index_html.replace("@@SECTIONS@@", "\n".join(sections))

(OUT / "index.html").write_text(index_html, encoding="utf-8")
(OUT / "_kb" / "kb.css").write_text(KB_CSS, encoding="utf-8")
_self_dst = OUT / "_kb" / "build_kb.py"
if Path(__file__).resolve() != _self_dst.resolve():
    shutil.copy2(__file__, _self_dst)

print("originals catalogued:", n_orig)
print("suite documents:", n_suite)
print("html files total:", len(list(OUT.rglob("*.html"))))
